from __future__ import annotations

import argparse
import math
import re
import string
from collections import Counter, defaultdict
from copy import copy
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook


DATA_RELATIVE_PATH = Path("\u8f6c\u5f55\u6570\u636e") / "Data.xlsx"
LLM_RELATIVE_PATH = Path("\u8f6c\u5f55\u6570\u636e") / "LLM\u8f6c\u5f55-new.xlsx"
TEMPLATE_RELATIVE_PATH = Path("\u8f6c\u5f55\u6570\u636e") / "\u6307\u6807\u8ba1\u7b97" / "field-all.xlsx"
OUTPUT_RELATIVE_PATH = Path("\u8f6c\u5f55\u6570\u636e") / "\u6307\u6807\u8ba1\u7b97" / "field-all-new.xlsx"

MODEL_SHEETS = [
    "glm",
    "gpt",
    "gemini",
    "qwen3.6-plus",
    "claude",
    "llama-4",
    "qwen3-vl-32b-instruct",
    "qwen2.5-vl-72b-instruct",
]

HEADERS = [
    "question_id",
    "LLM_answer_classifer",
    "LLM_answer_noun",
    "people_answer_calssifer",
    "LLM_answer_classifer_type",
    "people_answer_noun_type",
    "MRR",
    "people_answer__classifer_most",
    "people_answer_noun_most",
    "right_wrong",
    "accuracy",
    "Bleu",
    "MRR",
    "JSD",
]

TASK_QUANTIFIER = "\u91cf\u8bcd\u4efb\u52a1"
TASK_NOUN = "\u540d\u8bcd\u4efb\u52a1"
TASK_FREE_DESCRIPTION = "\u81ea\u7531\u63cf\u8ff0\u4efb\u52a1"

GROUPS = {
    "data_quantifier": {
        "label": "Data \u91cf\u8bcd\u4efb\u52a1",
        "task": TASK_QUANTIFIER,
        "kind": "quantifier",
        "rows": 180,
        "bleu": False,
        "jsd": True,
    },
    "data_noun": {
        "label": "Data \u540d\u8bcd\u4efb\u52a1",
        "task": TASK_NOUN,
        "kind": "noun",
        "rows": 120,
        "bleu": True,
        "jsd": False,
    },
    "free_quantifier": {
        "label": "Data \u81ea\u7531\u63cf\u8ff0-\u91cf\u8bcd",
        "task": TASK_FREE_DESCRIPTION,
        "kind": "quantifier",
        "rows": 60,
        "bleu": True,
        "jsd": True,
    },
    "free_noun": {
        "label": "Data \u81ea\u7531\u63cf\u8ff0-\u540d\u8bcd",
        "task": TASK_FREE_DESCRIPTION,
        "kind": "noun",
        "rows": 60,
        "bleu": True,
        "jsd": False,
    },
}

DATA_QUESTION_ID_IDX = 1
DATA_ANSWER_TEXT_IDX = 2
DATA_TASK_IDX = 3
DATA_QUANTIFIER_IDX = 4
DATA_NOUN_TYPE_IDX = 7
DATA_QUANTIFIER_TYPE_IDX = 6

PUNCTUATION_TABLE = str.maketrans(
    "",
    "",
    string.punctuation + "\u3002\uff0c\uff1f\uff01\uff1a\uff1b\u201c\u201d\u2018\u2019\u3001\uff08\uff09\u3010\u3011\u300a\u300b",
)


@dataclass
class AnswerRow:
    question_id: str
    answer_text: str
    task: str
    quantifier: str
    quantifier_type: str
    noun_type: str


@dataclass
class OutputRow:
    group_key: str
    question_id: str
    llm_quantifier: str = ""
    llm_noun_type: str = ""
    people_quantifier_distribution: str = ""
    llm_quantifier_type: str = ""
    people_noun_type_distribution: str = ""
    row_mrr: float | None = None
    people_quantifier_type_most: str = ""
    people_noun_type_most: str = ""
    right_wrong: int | None = None
    accuracy: float | None = None
    bleu: float | None = None
    group_mrr: float | None = None
    jsd: float | None = None

    def values(self) -> list[Any]:
        return [
            self.question_id,
            self.llm_quantifier or None,
            self.llm_noun_type or None,
            self.people_quantifier_distribution or None,
            self.llm_quantifier_type or None,
            self.people_noun_type_distribution or None,
            self.row_mrr,
            self.people_quantifier_type_most or None,
            self.people_noun_type_most or None,
            self.right_wrong,
            self.accuracy,
            self.bleu,
            self.group_mrr,
            self.jsd,
        ]


def clean_cell(value: object) -> str:
    if value is None:
        return ""
    return str(value).replace("_x000d_", "").replace("\r", "").replace("\n", "").strip()


def read_answer_row(row: tuple[Any, ...]) -> AnswerRow:
    return AnswerRow(
        question_id=clean_cell(row[DATA_QUESTION_ID_IDX]),
        answer_text=clean_cell(row[DATA_ANSWER_TEXT_IDX]),
        task=clean_cell(row[DATA_TASK_IDX]),
        quantifier=clean_cell(row[DATA_QUANTIFIER_IDX]),
        quantifier_type=clean_cell(row[DATA_QUANTIFIER_TYPE_IDX]),
        noun_type=clean_cell(row[DATA_NOUN_TYPE_IDX]),
    )


def distribution_text(counter: Counter[str]) -> str:
    return "; ".join(f"{key}:{count}" for key, count in counter.most_common())


def most_values(counter: Counter[str]) -> set[str]:
    if not counter:
        return set()
    max_count = max(counter.values())
    return {key for key, count in counter.items() if count == max_count}


def most_values_text(counter: Counter[str]) -> str:
    values = most_values(counter)
    if not values:
        return ""
    return ";".join(key for key in counter if key in values)


def reciprocal_rank(counter: Counter[str], answer: str) -> float | None:
    answer = clean_cell(answer)
    if not counter or not answer:
        return None
    if answer not in counter:
        return 0.0

    rank = 1
    previous_count: int | None = None
    for _, count in counter.most_common():
        if previous_count is None:
            previous_count = count
        elif count != previous_count:
            rank += 1
            previous_count = count
        if count == counter[answer]:
            return round(1 / rank, 6)
    return 0.0


def normalize_distribution(counts: Counter[str]) -> dict[str, float]:
    total = sum(counts.values())
    if total <= 0:
        return {}
    return {key: value / total for key, value in counts.items()}


def kl_divergence(p_dist: dict[str, float], q_dist: dict[str, float]) -> float:
    value = 0.0
    for key, p_value in p_dist.items():
        if p_value <= 0:
            continue
        q_value = q_dist.get(key, 0.0)
        if q_value <= 0:
            continue
        value += p_value * math.log2(p_value / q_value)
    return value


def jsd_from_counter_and_answer(counter: Counter[str], answer: str) -> float | None:
    answer = clean_cell(answer)
    if not counter or not answer:
        return None
    people_dist = normalize_distribution(counter)
    model_dist = {answer: 1.0}
    support = set(people_dist) | set(model_dist)
    midpoint = {
        key: (people_dist.get(key, 0.0) + model_dist.get(key, 0.0)) / 2
        for key in support
    }
    jsd = 0.5 * kl_divergence(people_dist, midpoint) + 0.5 * kl_divergence(model_dist, midpoint)
    return round(jsd, 6)


def tokenize_zh_chars(text: str) -> list[str]:
    normalized = clean_cell(text).translate(PUNCTUATION_TABLE)
    normalized = re.sub(r"\s+", "", normalized)
    return list(normalized)


def ngrams(tokens: list[str], n: int) -> Counter[tuple[str, ...]]:
    if len(tokens) < n:
        return Counter()
    return Counter(tuple(tokens[index : index + n]) for index in range(len(tokens) - n + 1))


def modified_precision_counts(
    candidate: list[str],
    references: list[list[str]],
    n: int,
) -> tuple[int, int]:
    candidate_ngrams = ngrams(candidate, n)
    total = sum(candidate_ngrams.values())
    if total == 0:
        return 0, 0

    max_ref_counts: Counter[tuple[str, ...]] = Counter()
    for reference in references:
        max_ref_counts |= ngrams(reference, n)
    clipped = sum(min(count, max_ref_counts[gram]) for gram, count in candidate_ngrams.items())
    return clipped, total


def closest_reference_length(candidate_len: int, reference_lengths: list[int]) -> int:
    return min(reference_lengths, key=lambda length: (abs(length - candidate_len), length))


def sentence_bleu_char(candidate_text: str, reference_texts: list[str], max_n: int = 4) -> float | None:
    candidate = tokenize_zh_chars(candidate_text)
    references = [tokens for text in reference_texts if (tokens := tokenize_zh_chars(text))]
    if not candidate or not references:
        return None

    candidate_len = len(candidate)
    reference_len = closest_reference_length(candidate_len, [len(reference) for reference in references])
    brevity_penalty = 1.0 if candidate_len > reference_len else math.exp(1 - reference_len / candidate_len)

    log_precision_sum = 0.0
    for n in range(1, max_n + 1):
        clipped, total = modified_precision_counts(candidate, references, n)
        precision = 1.0 if total == 0 else (clipped + 1) / (total + 1)
        log_precision_sum += math.log(precision)
    return round(brevity_penalty * math.exp(log_precision_sum / max_n), 6)


def load_people_data(data_path: Path) -> dict[str, dict[str, Any]]:
    workbook = load_workbook(data_path, read_only=True, data_only=True)
    try:
        data: dict[str, dict[str, Any]] = {
            group_key: {
                "quantifier": defaultdict(Counter),
                "quantifier_type": defaultdict(Counter),
                "noun_type": defaultdict(Counter),
                "texts": defaultdict(list),
            }
            for group_key in GROUPS
        }
        for row in workbook["people"].iter_rows(min_row=2, values_only=True):
            answer = read_answer_row(row)
            if not answer.question_id:
                continue
            for group_key, group in GROUPS.items():
                if answer.task != group["task"]:
                    continue
                if answer.quantifier:
                    data[group_key]["quantifier"][answer.question_id][answer.quantifier] += 1
                if answer.quantifier_type:
                    data[group_key]["quantifier_type"][answer.question_id][answer.quantifier_type] += 1
                if answer.noun_type:
                    data[group_key]["noun_type"][answer.question_id][answer.noun_type] += 1
                if answer.answer_text:
                    data[group_key]["texts"][answer.question_id].append(answer.answer_text)
        return data
    finally:
        workbook.close()


def load_llm_rows(llm_path: Path) -> dict[str, dict[str, list[AnswerRow]]]:
    workbook = load_workbook(llm_path, read_only=True, data_only=True)
    try:
        model_rows: dict[str, dict[str, list[AnswerRow]]] = {
            sheet_name: {group_key: [] for group_key in GROUPS}
            for sheet_name in MODEL_SHEETS
        }
        for sheet_name in MODEL_SHEETS:
            worksheet = workbook[sheet_name]
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                answer = read_answer_row(row)
                if not answer.question_id:
                    continue
                for group_key, group in GROUPS.items():
                    if answer.task == group["task"]:
                        # The free-description row is intentionally used twice:
                        # once for quantifier metrics and once for noun metrics.
                        model_rows[sheet_name][group_key].append(answer)
        return model_rows
    finally:
        workbook.close()


def build_output_rows(
    sheet_name: str,
    rows_by_group: dict[str, list[AnswerRow]],
    people_data: dict[str, dict[str, Any]],
) -> tuple[list[OutputRow], dict[str, dict[str, int | float]]]:
    output_rows: list[OutputRow] = []
    stats: dict[str, dict[str, int | float]] = {}

    for group_key, group in GROUPS.items():
        answer_rows = rows_by_group[group_key]
        group_output: list[OutputRow] = []
        row_mrr_values: list[float] = []
        right_wrong_values: list[int] = []
        missing_people = 0
        missing_model = 0

        for answer in answer_rows:
            if group["kind"] == "quantifier":
                content_counter = people_data[group_key]["quantifier"].get(answer.question_id, Counter())
                type_counter = people_data[group_key]["quantifier_type"].get(answer.question_id, Counter())
                row_mrr = reciprocal_rank(content_counter, answer.quantifier)
                most_types = most_values(type_counter)
                right_wrong = 1 if answer.quantifier_type and answer.quantifier_type in most_types else 0
                bleu = sentence_bleu_char(
                    answer.answer_text,
                    people_data[group_key]["texts"].get(answer.question_id, []),
                ) if group["bleu"] else None
                jsd = jsd_from_counter_and_answer(content_counter, answer.quantifier) if group["jsd"] else None
                output = OutputRow(
                    group_key=group_key,
                    question_id=answer.question_id,
                    llm_quantifier=answer.quantifier,
                    people_quantifier_distribution=distribution_text(content_counter),
                    llm_quantifier_type=answer.quantifier_type,
                    row_mrr=row_mrr,
                    people_quantifier_type_most=most_values_text(type_counter),
                    right_wrong=right_wrong if type_counter and answer.quantifier_type else None,
                    bleu=bleu,
                    jsd=jsd,
                )
                if not content_counter or not type_counter:
                    missing_people += 1
                if not answer.quantifier or not answer.quantifier_type:
                    missing_model += 1
            else:
                content_counter = people_data[group_key]["noun_type"].get(answer.question_id, Counter())
                row_mrr = reciprocal_rank(content_counter, answer.noun_type)
                most_noun_types = most_values(content_counter)
                right_wrong = 1 if answer.noun_type and answer.noun_type in most_noun_types else 0
                bleu = sentence_bleu_char(
                    answer.answer_text,
                    people_data[group_key]["texts"].get(answer.question_id, []),
                ) if group["bleu"] else None
                output = OutputRow(
                    group_key=group_key,
                    question_id=answer.question_id,
                    llm_noun_type=answer.noun_type,
                    people_noun_type_distribution=distribution_text(content_counter),
                    row_mrr=row_mrr,
                    people_noun_type_most=most_values_text(content_counter),
                    right_wrong=right_wrong if content_counter and answer.noun_type else None,
                    bleu=bleu,
                )
                if not content_counter:
                    missing_people += 1
                if not answer.noun_type:
                    missing_model += 1

            if output.row_mrr is not None:
                row_mrr_values.append(output.row_mrr)
            if output.right_wrong is not None:
                right_wrong_values.append(output.right_wrong)
            group_output.append(output)

        group_mrr = round(sum(row_mrr_values) / len(row_mrr_values), 6) if row_mrr_values else None
        accuracy = round(sum(right_wrong_values) / len(right_wrong_values), 6) if right_wrong_values else None
        for output in group_output:
            output.group_mrr = group_mrr
            output.accuracy = accuracy
        output_rows.extend(group_output)
        stats[group_key] = {
            "rows": len(answer_rows),
            "expected_rows": int(group["rows"]),
            "missing_people": missing_people,
            "missing_model": missing_model,
            "mean_mrr": group_mrr or 0.0,
            "accuracy": accuracy or 0.0,
            "bleu_nonblank": sum(1 for row in group_output if row.bleu is not None),
            "jsd_nonblank": sum(1 for row in group_output if row.jsd is not None),
        }

    if sheet_name and len(output_rows) != 420:
        stats["_sheet"] = {"rows": len(output_rows)}
    return output_rows, stats


def copy_sheet_format(template_ws, output_ws) -> None:
    output_ws.freeze_panes = template_ws.freeze_panes
    for col_idx in range(1, len(HEADERS) + 1):
        letter = template_ws.cell(row=1, column=col_idx).column_letter
        output_ws.column_dimensions[letter].width = template_ws.column_dimensions[letter].width

        source = template_ws.cell(row=1, column=col_idx)
        target = output_ws.cell(row=1, column=col_idx, value=HEADERS[col_idx - 1])
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.alignment:
            target.alignment = copy(source.alignment)
        if source.font:
            target.font = copy(source.font)
        if source.fill:
            target.fill = copy(source.fill)
        if source.border:
            target.border = copy(source.border)


def write_output_workbook(
    output_path: Path,
    template_path: Path,
    rows_by_sheet: dict[str, list[OutputRow]],
) -> None:
    template_workbook = load_workbook(template_path)
    try:
        workbook = Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)

        for sheet_name in MODEL_SHEETS:
            worksheet = workbook.create_sheet(sheet_name)
            template_ws = template_workbook[sheet_name] if sheet_name in template_workbook.sheetnames else template_workbook[MODEL_SHEETS[0]]
            copy_sheet_format(template_ws, worksheet)
            for row in rows_by_sheet[sheet_name]:
                worksheet.append(row.values())
        workbook.save(output_path)
    finally:
        template_workbook.close()


def print_stats(stats_by_sheet: dict[str, dict[str, dict[str, int | float]]]) -> None:
    for sheet_name in MODEL_SHEETS:
        print(f"\n{sheet_name}")
        for group_key, group in GROUPS.items():
            stats = stats_by_sheet[sheet_name][group_key]
            print(
                f"  {group['label']}: "
                f"rows={stats['rows']}, expected={stats['expected_rows']}, "
                f"missing_model={stats['missing_model']}, missing_people={stats['missing_people']}, "
                f"mean_mrr={stats['mean_mrr']}, accuracy={stats['accuracy']}, "
                f"bleu_nonblank={stats['bleu_nonblank']}, jsd_nonblank={stats['jsd_nonblank']}"
            )


def main() -> int:
    parser = argparse.ArgumentParser(description="Build field-all-new.xlsx metrics from LLM transcription new data.")
    parser.add_argument("--data", type=Path, default=DATA_RELATIVE_PATH)
    parser.add_argument("--llm", type=Path, default=LLM_RELATIVE_PATH)
    parser.add_argument("--template", type=Path, default=TEMPLATE_RELATIVE_PATH)
    parser.add_argument("--output", type=Path, default=OUTPUT_RELATIVE_PATH)
    parser.add_argument("--write", action="store_true", help="Write the output workbook. Without this, only print dry-run stats.")
    args = parser.parse_args()

    people_data = load_people_data(args.data)
    llm_rows = load_llm_rows(args.llm)

    rows_by_sheet: dict[str, list[OutputRow]] = {}
    stats_by_sheet: dict[str, dict[str, dict[str, int | float]]] = {}
    for sheet_name in MODEL_SHEETS:
        rows, stats = build_output_rows(sheet_name, llm_rows[sheet_name], people_data)
        rows_by_sheet[sheet_name] = rows
        stats_by_sheet[sheet_name] = stats

    print_stats(stats_by_sheet)

    if args.write:
        write_output_workbook(args.output, args.template, rows_by_sheet)
        print(f"\nwritten: {args.output}")
    else:
        print("\ndry-run only; pass --write to update the workbook")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
