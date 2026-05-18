from __future__ import annotations

import argparse
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook


ROOT_DIR = Path(__file__).resolve().parents[1]
ANSWER_DIR = ROOT_DIR / "LLM_answer_final"
DATA_DIR = ROOT_DIR / "\u8f6c\u5f55\u6570\u636e"

MATERIAL_PATH = DATA_DIR / "\u5b9e\u9a8c\u6750\u65991.xlsx"
TEMPLATE_PATH = DATA_DIR / "LLM\u8f6c\u5f55-new.xlsx"
TARGET_PATH = DATA_DIR / "LLM\u8f6c\u5f55-final.xlsx"
DATA_PATH = DATA_DIR / "Data.xlsx"
DEFAULT_REPORT_PATH = ROOT_DIR / "outputs" / "llm_transcription_final_report.xlsx"

COLUMNS = {
    "test_id": "test_id",
    "question_id": "question_id",
    "answer_text": "\u56de\u7b54\u6587\u672c",
    "condition": "\u5b9e\u9a8c\u6761\u4ef6",
    "quantifier": "\u91cf\u8bcd\uff08\u5185\u5bb9\uff09",
    "noun": "\u540d\u8bcd\uff08\u5185\u5bb9\uff09",
    "quantifier_type": "\u91cf\u8bcd\u7c7b\u578b",
    "noun_type": "\u540d\u8bcd\u7c7b\u578b",
    "item_time": "\u5355\u9898\u65f6\u95f4",
    "total_time": "\u603b\u65f6\u95f4",
}

MODEL_CONFIGS = [
    {"file_model": "glm-4.6v", "sheet": "glm", "test_id": "glm-4.6v"},
    {"file_model": "gpt-5.4", "sheet": "gpt", "test_id": "gpt-5.4"},
    {
        "file_model": "gemini-3.1-pro-preview",
        "sheet": "gemini",
        "test_id": "gemini-3.1-pro-preview",
    },
    {"file_model": "qwen3.6-plus", "sheet": "qwen3.6-plus", "test_id": "qwen3.6-plus"},
    {
        "file_model": "claude-sonnet-4-6",
        "sheet": "claude",
        "test_id": "claude-sonnet-4-6",
    },
    {"file_model": "llama-4-scout", "sheet": "llama-4", "test_id": "llama-4-scout"},
    {
        "file_model": "qwen3-vl-32b-instruct",
        "sheet": "qwen3-vl-32b-instruct",
        "test_id": "qwen3-vl-32b-instruct",
    },
    {
        "file_model": "qwen2.5-vl-72b-instruct",
        "sheet": "qwen2.5-vl-72b-instruct",
        "test_id": "qwen2.5-vl-72b-instruct",
    },
]

DATA_SHEETS = [config["sheet"] for config in MODEL_CONFIGS]

QUANTIFIERS = (
    "\u4e2a",
    "\u53ea",
    "\u4f4d",
    "\u540d",
    "\u8f86",
    "\u67b6",
    "\u8258",
    "\u682a",
    "\u68f5",
    "\u6735",
    "\u4ef6",
    "\u5f20",
    "\u6761",
    "\u628a",
    "\u95f4",
    "\u5ea7",
    "\u5bb6",
    "\u53f0",
    "\u672c",
    "\u5339",
    "\u5934",
    "\u53cc",
    "\u7247",
    "\u5757",
    "\u74f6",
    "\u676f",
    "\u7897",
    "\u76d2",
    "\u652f",
    "\u6839",
    "\u9876",
    "\u6247",
    "\u9762",
    "\u5c01",
    "\u5e45",
    "\u9897",
    "\u7c92",
    "\u76cf",
    "\u4efd",
    "\u53e3",
    "\u573a",
    "\u5957",
    "\u5bf9",
    "\u675f",
    "\u680b",
    "\u6b3e",
    "\u76d8",
    "\u79cd",
    "\u7ec4",
    "\u7fa4",
    "\u9053",
    "\u6392",
    "\u90e8",
)

QUESTION_ID_RE = re.compile(r"(rd(?:s|u|d|f|n)?[-_]\d+)\s*$")
ANSWER_HEADER_RE = re.compile(r"^\[(\d+)/(\d+)]\s+\u6587\u4ef6:\s+(.+?)\s*$")
BLOCK_SEPARATOR_RE = re.compile(r"^-{10,}\s*$")
ORIGINAL_OUTPUT_RE = re.compile(
    r"\u539f\u59cb\u8f93\u51fa:\s*(.*?)\s*\|\s*\u5355\u9898\u65f6\u95f4:\s*([-+]?\d+(?:\.\d+)?)\s*\u79d2",
    re.S,
)
IDENT_RE = re.compile(
    r"^LLM_[^_]+_(result_data(?:_special|_free|_downWord|_upWord|_normalWord)?)_"
    r"(\d+)_(rd(?:s|u|d|f|n)?[-_]\d+)$"
)
NUMERAL_CHARS = set("\u4e00\u4e24\u4e8c\u4e09\u56db\u4e94\u516d\u4e03\u516b\u4e5d\u53410123456789")
TRIM_CHARS = " \t\r\n\u3000\u3002\uff01\uff1f!?\uff0c\uff1b\u3001\"'\u201c\u201d\u2018\u2019\uff08\uff09()[]\u3010\u3011"
NOUN_SPLIT_RE = re.compile(r"[\n\r\u3002\uff01\uff1f!?\uff0c\uff1b\u3001]")


@dataclass
class AnswerRecord:
    index: int
    total: int
    identifier: str
    question_id: str
    answer_text: str
    item_time: float | None


@dataclass
class OutputRow:
    test_id: str
    question_id: str
    answer_text: str
    condition: str
    quantifier: str
    noun: str
    quantifier_type: str
    noun_type: str
    item_time: float | None
    total_time: float


def clean_cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("_x000d_", "").replace("\r", "").replace("\n", "").strip()


def resolve_path(value: str | None, default: Path) -> Path:
    if not value:
        return default
    path = Path(value)
    return path if path.is_absolute() else ROOT_DIR / path


def header_map(worksheet) -> dict[str, int]:
    result: dict[str, int] = {}
    for cell in worksheet[1]:
        name = clean_cell(cell.value)
        if name and name not in result:
            result[name] = cell.column
    return result


def require_columns(worksheet, names: list[str]) -> dict[str, int]:
    mapping = header_map(worksheet)
    missing = [name for name in names if name not in mapping]
    if missing:
        raise ValueError(f"Sheet {worksheet.title} missing columns: {missing}")
    return {name: mapping[name] for name in names}


def load_material_conditions(path: Path) -> dict[str, str]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    cols = require_columns(ws, [COLUMNS["question_id"], COLUMNS["condition"]])
    result: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2):
        question_id = clean_cell(row[cols[COLUMNS["question_id"]] - 1].value)
        condition = clean_cell(row[cols[COLUMNS["condition"]] - 1].value)
        if question_id:
            result[question_id] = condition
    wb.close()
    return result


def load_answer_fallback_lookup(path: Path) -> dict[str, dict[str, str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    required = [COLUMNS["question_id"], COLUMNS["answer_text"]]
    result: dict[str, dict[str, str]] = {}
    for sheet_name in DATA_SHEETS:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Data.xlsx missing sheet: {sheet_name}")
        ws = wb[sheet_name]
        cols = require_columns(ws, required)
        sheet_lookup: dict[str, str] = {}
        for row in ws.iter_rows(min_row=2):
            question_id = clean_cell(row[cols[COLUMNS["question_id"]] - 1].value)
            answer_text = clean_cell(row[cols[COLUMNS["answer_text"]] - 1].value)
            if question_id and answer_text:
                sheet_lookup[question_id] = answer_text
        result[sheet_name] = sheet_lookup
    wb.close()
    return result


def collect_global_type_lookup(
    path: Path,
    content_column: str,
    type_column: str,
) -> tuple[dict[str, str], dict[str, Counter[str]]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    counts_by_content: dict[str, Counter[str]] = defaultdict(Counter)
    for sheet_name in DATA_SHEETS:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Data.xlsx missing sheet: {sheet_name}")
        ws = wb[sheet_name]
        cols = require_columns(ws, [content_column, type_column])
        for row in ws.iter_rows(min_row=2):
            content = clean_cell(row[cols[content_column] - 1].value)
            type_value = clean_cell(row[cols[type_column] - 1].value)
            if content and type_value:
                counts_by_content[content][type_value] += 1
    wb.close()

    lookup: dict[str, str] = {}
    conflicts: dict[str, Counter[str]] = {}
    for content, counts in counts_by_content.items():
        if len(counts) == 1:
            lookup[content] = next(iter(counts))
        else:
            conflicts[content] = counts
    return lookup, conflicts


def parse_answer_file(path: Path) -> list[AnswerRecord]:
    text = path.read_text(encoding="utf-8-sig")
    records: list[AnswerRecord] = []
    current: dict[str, Any] | None = None
    block_lines: list[str] = []

    def finish_current() -> None:
        nonlocal current, block_lines
        if not current:
            return
        block = "\n".join(block_lines)
        output_match = ORIGINAL_OUTPUT_RE.search(block)
        answer_text = output_match.group(1).strip() if output_match else ""
        item_time = float(output_match.group(2)) if output_match else None
        records.append(
            AnswerRecord(
                index=current["index"],
                total=current["total"],
                identifier=current["identifier"],
                question_id=current["question_id"],
                answer_text=answer_text,
                item_time=item_time,
            )
        )
        current = None
        block_lines = []

    for raw_line in text.splitlines():
        line = raw_line.rstrip("\n")
        header_match = ANSWER_HEADER_RE.match(line.strip())
        if header_match:
            finish_current()
            identifier = header_match.group(3).strip()
            qid_match = QUESTION_ID_RE.search(identifier)
            current = {
                "index": int(header_match.group(1)),
                "total": int(header_match.group(2)),
                "identifier": identifier,
                "question_id": qid_match.group(1) if qid_match else "",
            }
            block_lines = []
            continue
        if not current:
            continue
        if BLOCK_SEPARATOR_RE.match(line.strip()):
            finish_current()
            continue
        block_lines.append(line)

    finish_current()
    return records


def trim_noun(text: str) -> str:
    value = text.strip(TRIM_CHARS)
    value = NOUN_SPLIT_RE.split(value, maxsplit=1)[0]
    return value.strip(TRIM_CHARS)


def extract_quantifier_noun(answer_text: str) -> tuple[str, str]:
    text = answer_text.replace("\\n", "\n").replace("\\r", "\n")
    candidates: list[tuple[str, str]] = []
    quantifiers = sorted(set(QUANTIFIERS), key=len, reverse=True)
    for index, char in enumerate(text):
        if char not in NUMERAL_CHARS:
            continue
        rest = text[index + 1 :]
        for quantifier in quantifiers:
            if rest.startswith(quantifier):
                noun = trim_noun(rest[len(quantifier) :])
                if noun:
                    candidates.append((quantifier, noun))
                break
    return candidates[-1] if candidates else ("", "")


def issue(
    config: dict[str, str],
    record: AnswerRecord | None,
    issue_type: str,
    detail: str,
    quantifier: str = "",
    noun: str = "",
    answer_text: str = "",
) -> dict[str, Any]:
    return {
        "model": config.get("file_model", ""),
        "sheet": config.get("sheet", ""),
        "row_index": record.index if record else "",
        "identifier": record.identifier if record else "",
        "question_id": record.question_id if record else "",
        "issue": issue_type,
        "detail": detail,
        "quantifier": quantifier,
        "noun": noun,
        "answer_text": answer_text or (record.answer_text if record else ""),
    }


def build_rows_for_model(
    config: dict[str, str],
    material_conditions: dict[str, str],
    answer_fallback_lookup: dict[str, dict[str, str]],
    quantifier_type_lookup: dict[str, str],
    noun_type_lookup: dict[str, str],
    quantifier_conflicts: dict[str, Counter[str]],
    noun_conflicts: dict[str, Counter[str]],
    report_rows: list[dict[str, Any]],
) -> tuple[list[OutputRow], dict[str, Any]]:
    answer_path = ANSWER_DIR / f"LLM_{config['file_model']}.txt"
    if not answer_path.exists():
        raise FileNotFoundError(answer_path)
    answers = parse_answer_file(answer_path)
    total_time = round(sum(record.item_time or 0 for record in answers), 2)

    stats = {
        "model": config["file_model"],
        "sheet": config["sheet"],
        "answers": len(answers),
        "times": sum(1 for record in answers if record.item_time is not None),
        "unique_question_ids": len({record.question_id for record in answers if record.question_id}),
        "duplicate_question_ids": len(answers) - len({record.question_id for record in answers}),
        "missing_material": 0,
        "answer_text_fallback": 0,
        "blank_answer_text": 0,
        "missing_extraction": 0,
        "missing_type": 0,
        "type_conflict": 0,
        "parse_count_issue": 0,
        "total_time": total_time,
    }
    if len(answers) != 360 or stats["times"] != 360 or stats["unique_question_ids"] != 360:
        stats["parse_count_issue"] = 1
        report_rows.append(
            issue(
                config,
                None,
                "parse_count_issue",
                f"answers={len(answers)}, times={stats['times']}, unique_question_ids={stats['unique_question_ids']}",
            )
        )

    rows: list[OutputRow] = []
    fallback_by_qid = answer_fallback_lookup.get(config["sheet"], {})
    for record in answers:
        condition = material_conditions.get(record.question_id, "")
        if not condition:
            stats["missing_material"] += 1
            report_rows.append(issue(config, record, "material_missing", "question_id not found in material"))

        answer_text = record.answer_text
        if not answer_text.strip():
            fallback_answer = fallback_by_qid.get(record.question_id, "")
            if fallback_answer:
                answer_text = fallback_answer
                stats["answer_text_fallback"] += 1
                report_rows.append(
                    issue(
                        config,
                        record,
                        "answer_text_fallback",
                        "blank original answer filled from Data.xlsx corresponding model sheet",
                        answer_text=answer_text,
                    )
                )
            else:
                stats["blank_answer_text"] += 1
                report_rows.append(issue(config, record, "blank_answer_text", "blank answer and no Data fallback"))

        quantifier, noun = extract_quantifier_noun(answer_text)
        if not quantifier or not noun:
            stats["missing_extraction"] += 1
            report_rows.append(
                issue(
                    config,
                    record,
                    "extraction_missing",
                    "could not parse quantifier/noun from final answer text",
                    quantifier=quantifier,
                    noun=noun,
                    answer_text=answer_text,
                )
            )

        quantifier_type = ""
        if quantifier:
            if quantifier in quantifier_conflicts:
                stats["type_conflict"] += 1
                report_rows.append(
                    issue(
                        config,
                        record,
                        "quantifier_type_conflict",
                        dict(quantifier_conflicts[quantifier]).__repr__(),
                        quantifier=quantifier,
                        noun=noun,
                    )
                )
            else:
                quantifier_type = quantifier_type_lookup.get(quantifier, "")

        noun_type = ""
        if noun:
            if noun in noun_conflicts:
                stats["type_conflict"] += 1
                report_rows.append(
                    issue(
                        config,
                        record,
                        "noun_type_conflict",
                        dict(noun_conflicts[noun]).__repr__(),
                        quantifier=quantifier,
                        noun=noun,
                    )
                )
            else:
                noun_type = noun_type_lookup.get(noun, "")

        if (quantifier and not quantifier_type) or (noun and not noun_type):
            stats["missing_type"] += 1
            report_rows.append(
                issue(
                    config,
                    record,
                    "type_missing",
                    "no unique global Data.xlsx type match for quantifier or noun",
                    quantifier=quantifier,
                    noun=noun,
                )
            )

        rows.append(
            OutputRow(
                test_id=config["test_id"],
                question_id=record.question_id,
                answer_text=answer_text,
                condition=condition,
                quantifier=quantifier,
                noun=noun,
                quantifier_type=quantifier_type,
                noun_type=noun_type,
                item_time=record.item_time,
                total_time=total_time,
            )
        )

    return rows, stats


def clear_sheet_data(worksheet) -> None:
    if worksheet.max_row >= 2:
        worksheet.delete_rows(2, worksheet.max_row - 1)


def write_rows_to_sheet(worksheet, rows: list[OutputRow]) -> None:
    cols = require_columns(worksheet, list(COLUMNS.values()))
    field_order = [
        ("test_id", "test_id"),
        ("question_id", "question_id"),
        ("answer_text", "answer_text"),
        ("condition", "condition"),
        ("quantifier", "quantifier"),
        ("noun", "noun"),
        ("quantifier_type", "quantifier_type"),
        ("noun_type", "noun_type"),
        ("item_time", "item_time"),
        ("total_time", "total_time"),
    ]
    for row_index, output_row in enumerate(rows, start=2):
        for attr_name, column_key in field_order:
            worksheet.cell(row=row_index, column=cols[COLUMNS[column_key]], value=getattr(output_row, attr_name))


def save_report(path: Path, stats_rows: list[dict[str, Any]], issue_rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "summary"
    if stats_rows:
        headers = list(stats_rows[0].keys())
        summary_ws.append(headers)
        for row in stats_rows:
            summary_ws.append([row.get(header, "") for header in headers])
    else:
        summary_ws.append(["message"])
        summary_ws.append(["no stats"])

    issue_ws = wb.create_sheet("issues")
    if issue_rows:
        headers = list(issue_rows[0].keys())
        issue_ws.append(headers)
        for row in issue_rows:
            issue_ws.append([row.get(header, "") for header in headers])
    else:
        issue_ws.append(["message"])
        issue_ws.append(["no issues"])
    wb.save(path)


def process(output_path: Path, report_path: Path, dry_run: bool) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    material_conditions = load_material_conditions(MATERIAL_PATH)
    answer_fallback_lookup = load_answer_fallback_lookup(DATA_PATH)
    quantifier_type_lookup, quantifier_conflicts = collect_global_type_lookup(
        DATA_PATH, COLUMNS["quantifier"], COLUMNS["quantifier_type"]
    )
    noun_type_lookup, noun_conflicts = collect_global_type_lookup(DATA_PATH, COLUMNS["noun"], COLUMNS["noun_type"])

    report_rows: list[dict[str, Any]] = []
    for content, counts in quantifier_conflicts.items():
        report_rows.append(
            {
                "model": "",
                "sheet": "",
                "row_index": "",
                "identifier": "",
                "question_id": "",
                "issue": "global_quantifier_type_conflict",
                "detail": dict(counts).__repr__(),
                "quantifier": content,
                "noun": "",
                "answer_text": "",
            }
        )
    for content, counts in noun_conflicts.items():
        report_rows.append(
            {
                "model": "",
                "sheet": "",
                "row_index": "",
                "identifier": "",
                "question_id": "",
                "issue": "global_noun_type_conflict",
                "detail": dict(counts).__repr__(),
                "quantifier": "",
                "noun": content,
                "answer_text": "",
            }
        )

    stats_rows: list[dict[str, Any]] = []
    rows_by_sheet: dict[str, list[OutputRow]] = {}
    for config in MODEL_CONFIGS:
        rows, stats = build_rows_for_model(
            config=config,
            material_conditions=material_conditions,
            answer_fallback_lookup=answer_fallback_lookup,
            quantifier_type_lookup=quantifier_type_lookup,
            noun_type_lookup=noun_type_lookup,
            quantifier_conflicts=quantifier_conflicts,
            noun_conflicts=noun_conflicts,
            report_rows=report_rows,
        )
        rows_by_sheet[config["sheet"]] = rows
        stats_rows.append(stats)

    if not dry_run:
        wb = load_workbook(TEMPLATE_PATH)
        for config in MODEL_CONFIGS:
            sheet_name = config["sheet"]
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"Template workbook missing sheet: {sheet_name}")
            ws = wb[sheet_name]
            clear_sheet_data(ws)
            write_rows_to_sheet(ws, rows_by_sheet[sheet_name])
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        save_report(report_path, stats_rows, report_rows)

    return stats_rows, report_rows


def print_summary(stats_rows: list[dict[str, Any]], report_rows: list[dict[str, Any]], dry_run: bool) -> None:
    mode = "DRY RUN" if dry_run else "WRITE"
    print(f"Mode: {mode}")
    print(
        "model | answers | times | unique_qids | missing_material | answer_text_fallback | "
        "blank_answer_text | missing_extraction | missing_type | type_conflict | total_time"
    )
    for row in stats_rows:
        print(
            f"{row['model']} | {row['answers']} | {row['times']} | {row['unique_question_ids']} | "
            f"{row['missing_material']} | {row['answer_text_fallback']} | {row['blank_answer_text']} | "
            f"{row['missing_extraction']} | {row['missing_type']} | {row['type_conflict']} | {row['total_time']}"
        )
    print(f"Total output rows: {sum(row['answers'] for row in stats_rows)}")
    print(f"Report issue rows: {len(report_rows)}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Fill final LLM transcription workbook from final answer files.")
    parser.add_argument("--dry-run", action="store_true", help="Parse and validate without saving workbooks.")
    parser.add_argument("--output", help="Output xlsx path. Defaults to LLM\u8f6c\u5f55-final.xlsx.")
    parser.add_argument("--report", help="Report xlsx path. Defaults to outputs/llm_transcription_final_report.xlsx.")
    args = parser.parse_args()

    output_path = resolve_path(args.output, TARGET_PATH)
    report_path = resolve_path(args.report, DEFAULT_REPORT_PATH)
    stats_rows, report_rows = process(output_path=output_path, report_path=report_path, dry_run=args.dry_run)
    print_summary(stats_rows, report_rows, dry_run=args.dry_run)
    if not args.dry_run:
        print(f"Saved workbook: {output_path}")
        print(f"Saved report: {report_path}")


if __name__ == "__main__":
    main()
