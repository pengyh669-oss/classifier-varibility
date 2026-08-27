from __future__ import annotations

import argparse
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from build_baseline_task1 import (
    clean_cell,
    jsd_from_counter_and_answer,
    mean,
    parse_distribution_text,
    reciprocal_rank,
)


INPUT_RELATIVE_PATH = (
    Path("\u8f6c\u5f55\u6570\u636e")
    / "\u6307\u6807\u8ba1\u7b97"
    / "field-all-final.xlsx"
)
OUTPUT_RELATIVE_PATH = (
    Path("\u8f6c\u5f55\u6570\u636e")
    / "\u6307\u6807\u8ba1\u7b97"
    / "baseline-task1-experiment2.xlsx"
)

DEFAULT_SHEET = "gpt"
TASK1_START_ROW = 2
TASK1_END_ROW = 181

QUESTION_ID_COL = 1
PEOPLE_DISTRIBUTION_COL = 4
PEOPLE_MOST_TYPE_COL = 8


@dataclass
class DetailRow:
    question_id: str
    question_suffix: int
    animacy_group: str
    people_distribution_text: str
    baseline_classifier: str
    baseline_classifier_type: str
    people_most_type: str
    rr: float
    right_wrong: int
    jsd: float

    def values(self) -> list[Any]:
        return [
            self.question_id,
            self.question_suffix,
            self.animacy_group,
            self.people_distribution_text,
            self.baseline_classifier,
            self.baseline_classifier_type,
            self.people_most_type,
            self.rr,
            self.right_wrong,
            self.jsd,
        ]


@dataclass
class SkippedRow:
    question_id: str
    row_number: int
    reason: str

    def values(self) -> list[Any]:
        return [self.question_id, self.row_number, self.reason]


def parse_question_suffix(question_id: str) -> int | None:
    match = re.search(r"(\d+)$", clean_cell(question_id))
    if not match:
        return None
    return int(match.group(1))


def baseline_for_suffix(suffix: int) -> tuple[str, str, str] | None:
    if 1 <= suffix <= 30:
        return "animate", "\u53ea", "\u7279\u6b8a\u91cf\u8bcd"
    if 31 <= suffix <= 60:
        return "inanimate", "\u4e2a", "\u666e\u901a\u91cf\u8bcd"
    return None


def split_type_set(value: object) -> set[str]:
    text = clean_cell(value)
    if not text:
        return set()
    return {
        clean_cell(part)
        for part in re.split(r"[;\uff1b]", text)
        if clean_cell(part)
    }


def build_rows(input_path: Path, sheet_name: str) -> tuple[list[DetailRow], list[SkippedRow]]:
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet_name]
        detail_rows: list[DetailRow] = []
        skipped_rows: list[SkippedRow] = []

        for row_number in range(TASK1_START_ROW, TASK1_END_ROW + 1):
            question_id = clean_cell(worksheet.cell(row=row_number, column=QUESTION_ID_COL).value)
            suffix = parse_question_suffix(question_id)
            if suffix is None:
                skipped_rows.append(SkippedRow(question_id, row_number, "cannot parse question_id suffix"))
                continue

            baseline = baseline_for_suffix(suffix)
            if baseline is None:
                skipped_rows.append(SkippedRow(question_id, row_number, "question_id suffix outside 1-60"))
                continue

            distribution_text = clean_cell(worksheet.cell(row=row_number, column=PEOPLE_DISTRIBUTION_COL).value)
            counts = parse_distribution_text(distribution_text)
            if not counts:
                skipped_rows.append(
                    SkippedRow(question_id, row_number, "empty or unparsable people_answer_calssifer")
                )
                continue

            people_most_type = clean_cell(worksheet.cell(row=row_number, column=PEOPLE_MOST_TYPE_COL).value)
            people_most_type_set = split_type_set(people_most_type)
            animacy_group, baseline_classifier, baseline_classifier_type = baseline
            rr = reciprocal_rank(counts, baseline_classifier)
            jsd = jsd_from_counter_and_answer(counts, baseline_classifier)
            if rr is None or jsd is None:
                skipped_rows.append(SkippedRow(question_id, row_number, "failed to calculate RR or JSD"))
                continue

            right_wrong = 1 if baseline_classifier_type in people_most_type_set else 0
            detail_rows.append(
                DetailRow(
                    question_id=question_id,
                    question_suffix=suffix,
                    animacy_group=animacy_group,
                    people_distribution_text=distribution_text,
                    baseline_classifier=baseline_classifier,
                    baseline_classifier_type=baseline_classifier_type,
                    people_most_type=people_most_type,
                    rr=rr,
                    right_wrong=right_wrong,
                    jsd=jsd,
                )
            )
        return detail_rows, skipped_rows
    finally:
        workbook.close()


def write_output(output_path: Path, detail_rows: list[DetailRow], skipped_rows: list[SkippedRow]) -> None:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "summary"
    detail = workbook.create_sheet("detail")
    skipped = workbook.create_sheet("skipped")

    summary_headers = [
        "baseline_name",
        "valid_items",
        "skipped_items",
        "animate_items",
        "inanimate_items",
        "MRR",
        "accuracy",
        "JSD",
    ]
    detail_headers = [
        "question_id",
        "question_suffix",
        "animacy_group",
        "people_answer_calssifer",
        "baseline_answer_classifier",
        "baseline_answer_classifier_type",
        "people_answer__classifer_most",
        "RR",
        "right_wrong",
        "JSD",
    ]
    skipped_headers = ["question_id", "row", "reason"]

    animate_items = sum(1 for row in detail_rows if row.animacy_group == "animate")
    inanimate_items = sum(1 for row in detail_rows if row.animacy_group == "inanimate")

    summary.append(summary_headers)
    summary.append(
        [
            "Animacy Rule Classifier",
            len(detail_rows),
            len(skipped_rows),
            animate_items,
            inanimate_items,
            mean([row.rr for row in detail_rows]),
            mean([row.right_wrong for row in detail_rows]),
            mean([row.jsd for row in detail_rows]),
        ]
    )

    detail.append(detail_headers)
    for row in detail_rows:
        detail.append(row.values())

    skipped.append(skipped_headers)
    for row in skipped_rows:
        skipped.append(row.values())

    for worksheet in (summary, detail, skipped):
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
        worksheet.freeze_panes = "A2"
        for column_cells in worksheet.columns:
            max_length = max(len(clean_cell(cell.value)) for cell in column_cells)
            worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 10), 48)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def print_summary(output_path: Path, detail_rows: list[DetailRow], skipped_rows: list[SkippedRow]) -> None:
    valid_items = len(detail_rows)
    skipped_items = len(skipped_rows)
    animate_items = sum(1 for row in detail_rows if row.animacy_group == "animate")
    inanimate_items = sum(1 for row in detail_rows if row.animacy_group == "inanimate")
    mrr = mean([row.rr for row in detail_rows])
    accuracy = mean([row.right_wrong for row in detail_rows])
    jsd = mean([row.jsd for row in detail_rows])
    skipped_question_ids = [row.question_id for row in skipped_rows]

    print(f"output_file={output_path}")
    print(f"valid_items={valid_items}")
    print(f"skipped_items={skipped_items}")
    print(f"animate_items={animate_items}")
    print(f"inanimate_items={inanimate_items}")
    print(f"MRR={mrr}")
    print(f"accuracy={accuracy}")
    print(f"JSD={jsd}")
    print(f"skipped_question_id_list={skipped_question_ids}")

    if valid_items != TASK1_END_ROW - TASK1_START_ROW + 1:
        print(f"warning: valid_items={valid_items}, expected={TASK1_END_ROW - TASK1_START_ROW + 1}")
    if animate_items != 90 or inanimate_items != 90:
        print(f"warning: animate_items={animate_items}, inanimate_items={inanimate_items}, expected=90/90")


def main() -> int:
    parser = argparse.ArgumentParser(description="Build Task 1 animacy-rule classifier baseline workbook.")
    parser.add_argument("--input", type=Path, default=INPUT_RELATIVE_PATH)
    parser.add_argument("--output", type=Path, default=OUTPUT_RELATIVE_PATH)
    parser.add_argument("--sheet", default=DEFAULT_SHEET)
    args = parser.parse_args()

    detail_rows, skipped_rows = build_rows(args.input, args.sheet)
    write_output(args.output, detail_rows, skipped_rows)
    print_summary(args.output, detail_rows, skipped_rows)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
