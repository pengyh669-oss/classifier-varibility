from __future__ import annotations

import argparse
from copy import copy
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from build_field_all_new_metrics import (
    GROUPS,
    MODEL_SHEETS,
    build_output_rows,
    load_llm_rows,
    load_people_data,
)


DATA_RELATIVE_PATH = Path("\u8f6c\u5f55\u6570\u636e") / "Data.xlsx"
LLM_RELATIVE_PATH = Path("\u8f6c\u5f55\u6570\u636e") / "LLM\u8f6c\u5f55-final.xlsx"
TEMPLATE_RELATIVE_PATH = (
    Path("\u8f6c\u5f55\u6570\u636e")
    / "\u6307\u6807\u8ba1\u7b97"
    / "\u7edf\u8ba1\u6307\u6807-new.xlsx"
)
OUTPUT_RELATIVE_PATH = (
    Path("\u8f6c\u5f55\u6570\u636e")
    / "\u6307\u6807\u8ba1\u7b97"
    / "\u7edf\u8ba1\u6307\u6807-final.xlsx"
)

MODEL_DISPLAY_TO_SHEET = {
    "glm": "glm",
    "gpt": "gpt",
    "gemini": "gemini",
    "qwen3.6-plus": "qwen3.6-plus",
    "claude": "claude",
    "llama-4": "llama-4",
    "qwen3-vl": "qwen3-vl-32b-instruct",
    "qwne2.5-vl": "qwen2.5-vl-72b-instruct",
}

HEADERS = [
    "\u6a21\u578b\u540d\u5b57",
    "\u603b\u4f53-\u91cf\u8bcd-MRR",
    "\u603b\u4f53-\u540d\u8bcd-MRR",
    "\u603b\u4f53-\u81ea\u7531\u91cf\u8bcd-MRR",
    "\u603b\u4f53-\u81ea\u7531\u540d\u8bcd-MRR",
    "\u603b\u4f53-\u91cf\u8bcd-accuracy",
    "\u603b\u4f53-\u540d\u8bcd-accuracy",
    "\u603b\u4f53-\u81ea\u7531\u91cf\u8bcd-accuracy",
    "\u603b\u4f53-\u81ea\u7531\u540d\u8bcd-accuracy",
    "\u603b\u4f53-\u540d\u8bcd-bleu",
    "\u603b\u4f53-\u81ea\u7531-bleu",
    "\u603b\u4f53-\u91cf\u8bcd-jsd",
    "\u603b\u4f53-\u81ea\u7531\u91cf\u8bcd-jsd",
]

METRIC_COLUMNS = [
    ("data_quantifier", "mrr"),
    ("data_noun", "mrr"),
    ("free_quantifier", "mrr"),
    ("free_noun", "mrr"),
    ("data_quantifier", "accuracy"),
    ("data_noun", "accuracy"),
    ("free_quantifier", "accuracy"),
    ("free_noun", "accuracy"),
    ("data_noun", "bleu"),
    ("free_quantifier", "bleu"),
    ("data_quantifier", "jsd"),
    ("free_quantifier", "jsd"),
]


def mean(values: list[float]) -> float | None:
    if not values:
        return None
    return round(sum(values) / len(values), 6)


def build_model_metrics(
    data_path: Path,
    llm_path: Path,
) -> tuple[dict[str, dict[str, dict[str, float | None]]], dict[str, dict[str, dict[str, Any]]]]:
    people_data = load_people_data(data_path)
    llm_rows = load_llm_rows(llm_path)
    metrics: dict[str, dict[str, dict[str, float | None]]] = {}
    report: dict[str, dict[str, dict[str, Any]]] = {}

    for sheet_name in MODEL_SHEETS:
        output_rows, stats = build_output_rows(sheet_name, llm_rows[sheet_name], people_data)
        metrics[sheet_name] = {}
        report[sheet_name] = stats

        rows_by_group = {group_key: [] for group_key in GROUPS}
        for row in output_rows:
            rows_by_group[row.group_key].append(row)

        for group_key, rows in rows_by_group.items():
            metrics[sheet_name][group_key] = {
                "mrr": stats[group_key]["mean_mrr"],
                "accuracy": stats[group_key]["accuracy"],
                "bleu": mean([row.bleu for row in rows if row.bleu is not None]),
                "jsd": mean([row.jsd for row in rows if row.jsd is not None]),
            }

    return metrics, report


def copy_template_shape(template_path: Path, worksheet) -> None:
    if not template_path.exists() or template_path.stat().st_size == 0:
        return

    template_workbook = load_workbook(template_path)
    try:
        template_ws = template_workbook[template_workbook.sheetnames[0]]
        worksheet.freeze_panes = template_ws.freeze_panes
        for col_idx in range(1, len(HEADERS) + 1):
            letter = worksheet.cell(row=1, column=col_idx).column_letter
            worksheet.column_dimensions[letter].width = template_ws.column_dimensions[letter].width
            source = template_ws.cell(row=1, column=col_idx)
            target = worksheet.cell(row=1, column=col_idx)
            if source.has_style:
                target._style = copy(source._style)
            target.number_format = source.number_format
            target.font = copy(source.font)
            target.fill = copy(source.fill)
            target.border = copy(source.border)
            target.alignment = copy(source.alignment)
            target.protection = copy(source.protection)
    finally:
        template_workbook.close()


def write_statistics_workbook(
    output_path: Path,
    template_path: Path,
    metrics: dict[str, dict[str, dict[str, float | None]]],
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "\u7edf\u8ba1\u6307\u6807-final"

    for col_idx, header in enumerate(HEADERS, start=1):
        worksheet.cell(row=1, column=col_idx, value=header)
    copy_template_shape(template_path, worksheet)

    for row_idx, (display_name, sheet_name) in enumerate(MODEL_DISPLAY_TO_SHEET.items(), start=2):
        worksheet.cell(row=row_idx, column=1, value=display_name)
        for offset, (group_key, metric_key) in enumerate(METRIC_COLUMNS, start=2):
            value = metrics[sheet_name][group_key][metric_key]
            worksheet.cell(row=row_idx, column=offset, value=value)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def print_report(
    metrics: dict[str, dict[str, dict[str, float | None]]],
    report: dict[str, dict[str, dict[str, Any]]],
) -> None:
    for display_name, sheet_name in MODEL_DISPLAY_TO_SHEET.items():
        print(f"\n{display_name} ({sheet_name})")
        for group_key, group in GROUPS.items():
            stats = report[sheet_name][group_key]
            values = metrics[sheet_name][group_key]
            print(
                f"  {group['label']}: rows={stats['rows']}, expected={stats['expected_rows']}, "
                f"missing_model={stats['missing_model']}, missing_people={stats['missing_people']}, "
                f"mrr={values['mrr']}, accuracy={values['accuracy']}, "
                f"bleu={values['bleu']}, jsd={values['jsd']}"
            )


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Calculate overall final statistics from LLM\u8f6c\u5f55-final.xlsx and Data.xlsx/people."
    )
    parser.add_argument("--data", type=Path, default=DATA_RELATIVE_PATH)
    parser.add_argument("--llm", type=Path, default=LLM_RELATIVE_PATH)
    parser.add_argument("--template", type=Path, default=TEMPLATE_RELATIVE_PATH)
    parser.add_argument("--output", type=Path, default=OUTPUT_RELATIVE_PATH)
    parser.add_argument("--write", action="store_true", help="Write \u7edf\u8ba1\u6307\u6807-final.xlsx.")
    args = parser.parse_args()

    metrics, report = build_model_metrics(args.data, args.llm)
    print_report(metrics, report)

    if args.write:
        write_statistics_workbook(args.output, args.template, metrics)
        print(f"\nwritten: {args.output}")
    else:
        print("\ndry-run only; pass --write to update the workbook")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
