from __future__ import annotations

import argparse
import math
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook


DATA_RELATIVE_PATH = Path("\u8f6c\u5f55\u6570\u636e") / "Data.xlsx"
FIELD_RELATIVE_PATHS = [
    Path("\u8f6c\u5f55\u6570\u636e") / "\u6307\u6807\u8ba1\u7b97" / "field-all.xlsx",
    Path("\u8f6c\u5f55\u6570\u636e") / "\u6307\u6807\u8ba1\u7b97" / "field-live.xlsx",
    Path("\u8f6c\u5f55\u6570\u636e") / "\u6307\u6807\u8ba1\u7b97" / "field-unlive.xlsx",
]

MODEL_SHEETS = [
    "glm",
    "gpt",
    "gemini",
    "qwen3.6-plus",
    "claude",
    "llama-4",
    "qwen3-vl-32b-instruct",
    "qwen2.5-vl-72b-instruct",
]

TASK_QUANTIFIER = "\u91cf\u8bcd\u4efb\u52a1"
TASK_FREE_DESCRIPTION = "\u81ea\u7531\u63cf\u8ff0\u4efb\u52a1"

FIELD_GROUPS = {
    "data_quantifier": {
        "label": "Data \u91cf\u8bcd\u4efb\u52a1",
        "data_task": TASK_QUANTIFIER,
    },
    "free_description_quantifier": {
        "label": "Data \u81ea\u7531\u63cf\u8ff0-\u91cf\u8bcd",
        "data_task": TASK_FREE_DESCRIPTION,
    },
}

FIELD_ROW_RANGES = {
    "field-all.xlsx": {
        "data_quantifier": (2, 181),
        "free_description_quantifier": (302, 361),
    },
    "field-live.xlsx": {
        "data_quantifier": (2, 91),
        "free_description_quantifier": (152, 181),
    },
    "field-unlive.xlsx": {
        "data_quantifier": (2, 91),
        "free_description_quantifier": (152, 181),
    },
}

QUESTION_ID_COL = 1
FIELD_LLM_QUANTIFIER_COL = 2
DATA_QUESTION_ID_IDX = 1
DATA_TASK_IDX = 3
DATA_QUANTIFIER_IDX = 4


def clean_cell(value: object) -> str:
    if value is None:
        return ""
    return str(value).replace("_x000d_", "").replace("\r", "").replace("\n", "").strip()


def normalize_distribution(counts: Counter[str]) -> dict[str, float]:
    total = sum(counts.values())
    if total <= 0:
        return {}
    return {key: value / total for key, value in counts.items()}


def kl_divergence(p_dist: dict[str, float], q_dist: dict[str, float]) -> float:
    value = 0.0
    for key, p_value in p_dist.items():
        if p_value <= 0:
            continue
        q_value = q_dist.get(key, 0.0)
        if q_value <= 0:
            continue
        value += p_value * math.log2(p_value / q_value)
    return value


def jsd_from_counter_and_answer(people_counts: Counter[str], model_answer: str) -> float | None:
    model_answer = clean_cell(model_answer)
    if not people_counts or not model_answer:
        return None

    people_dist = normalize_distribution(people_counts)
    model_dist = {model_answer: 1.0}
    support = set(people_dist) | set(model_dist)
    midpoint = {
        key: (people_dist.get(key, 0.0) + model_dist.get(key, 0.0)) / 2
        for key in support
    }
    jsd = 0.5 * kl_divergence(people_dist, midpoint) + 0.5 * kl_divergence(model_dist, midpoint)
    return round(jsd, 6)


def load_people_quantifier_counts(data_path: Path) -> dict[str, dict[str, Counter[str]]]:
    workbook = load_workbook(data_path, read_only=True, data_only=True)
    try:
        people_counts: dict[str, dict[str, Counter[str]]] = {
            group_key: defaultdict(Counter) for group_key in FIELD_GROUPS
        }
        worksheet = workbook["people"]
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            question_id = clean_cell(row[DATA_QUESTION_ID_IDX])
            task = clean_cell(row[DATA_TASK_IDX])
            quantifier = clean_cell(row[DATA_QUANTIFIER_IDX])
            if not question_id or not quantifier:
                continue
            for group_key, group in FIELD_GROUPS.items():
                if task == group["data_task"]:
                    people_counts[group_key][question_id][quantifier] += 1
        return {
            group_key: dict(counts_by_question)
            for group_key, counts_by_question in people_counts.items()
        }
    finally:
        workbook.close()


def load_model_quantifier_answers(data_path: Path) -> dict[str, dict[str, dict[str, str]]]:
    workbook = load_workbook(data_path, read_only=True, data_only=True)
    try:
        answers: dict[str, dict[str, dict[str, str]]] = {
            group_key: {sheet_name: {} for sheet_name in MODEL_SHEETS}
            for group_key in FIELD_GROUPS
        }
        for sheet_name in MODEL_SHEETS:
            worksheet = workbook[sheet_name]
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                question_id = clean_cell(row[DATA_QUESTION_ID_IDX])
                task = clean_cell(row[DATA_TASK_IDX])
                quantifier = clean_cell(row[DATA_QUANTIFIER_IDX])
                if not question_id or not quantifier:
                    continue
                for group_key, group in FIELD_GROUPS.items():
                    if task == group["data_task"]:
                        answers[group_key][sheet_name][question_id] = quantifier
        return answers
    finally:
        workbook.close()


def get_or_create_jsd_column(worksheet) -> int:
    header = [
        clean_cell(worksheet.cell(row=1, column=col).value)
        for col in range(1, worksheet.max_column + 1)
    ]
    if "JSD" in header:
        return header.index("JSD") + 1

    jsd_col = worksheet.max_column + 1
    worksheet.cell(row=1, column=jsd_col, value="JSD")
    return jsd_col


def field_ranges_for_path(field_path: Path) -> dict[str, tuple[int, int]]:
    try:
        return FIELD_ROW_RANGES[field_path.name]
    except KeyError as exc:
        known = ", ".join(sorted(FIELD_ROW_RANGES))
        raise ValueError(f"No configured row ranges for {field_path.name}. Known files: {known}") from exc


def calculate_field_jsd(
    field_path: Path,
    people_counts: dict[str, dict[str, Counter[str]]],
    model_answers: dict[str, dict[str, dict[str, str]]],
) -> dict[str, dict[str, dict[str, int | float]]]:
    workbook = load_workbook(field_path, read_only=True, data_only=True)
    try:
        ranges = field_ranges_for_path(field_path)
        stats: dict[str, dict[str, dict[str, int | float]]] = {
            group_key: {
                sheet_name: {
                    "target_rows": 0,
                    "calculated": 0,
                    "missing_people": 0,
                    "missing_model": 0,
                    "mean_jsd": 0.0,
                }
                for sheet_name in MODEL_SHEETS
            }
            for group_key in FIELD_GROUPS
        }

        for sheet_name in MODEL_SHEETS:
            worksheet = workbook[sheet_name]
            for group_key, (start_row, end_row) in ranges.items():
                scores: list[float] = []
                sheet_stats = stats[group_key][sheet_name]
                for row_idx in range(start_row, end_row + 1):
                    question_id = clean_cell(worksheet.cell(row=row_idx, column=QUESTION_ID_COL).value)
                    if not question_id:
                        continue
                    sheet_stats["target_rows"] += 1
                    answer = model_answers[group_key][sheet_name].get(question_id, "")
                    if not answer:
                        sheet_stats["missing_model"] += 1
                        continue
                    counts = people_counts[group_key].get(question_id, Counter())
                    if not counts:
                        sheet_stats["missing_people"] += 1
                        continue
                    score = jsd_from_counter_and_answer(counts, answer)
                    if score is None:
                        continue
                    scores.append(score)
                    sheet_stats["calculated"] += 1
                if scores:
                    sheet_stats["mean_jsd"] = round(sum(scores) / len(scores), 6)
        return stats
    finally:
        workbook.close()


def write_field_jsd(
    field_path: Path,
    people_counts: dict[str, dict[str, Counter[str]]],
    model_answers: dict[str, dict[str, dict[str, str]]],
) -> dict[str, dict[str, dict[str, int | float]]]:
    workbook = load_workbook(field_path)
    try:
        ranges = field_ranges_for_path(field_path)
        stats: dict[str, dict[str, dict[str, int | float]]] = {
            group_key: {
                sheet_name: {
                    "target_rows": 0,
                    "written": 0,
                    "missing_people": 0,
                    "missing_model": 0,
                    "mean_jsd": 0.0,
                }
                for sheet_name in MODEL_SHEETS
            }
            for group_key in FIELD_GROUPS
        }

        for sheet_name in MODEL_SHEETS:
            worksheet = workbook[sheet_name]
            jsd_col = get_or_create_jsd_column(worksheet)
            for group_key, (start_row, end_row) in ranges.items():
                scores: list[float] = []
                sheet_stats = stats[group_key][sheet_name]
                for row_idx in range(start_row, end_row + 1):
                    question_id = clean_cell(worksheet.cell(row=row_idx, column=QUESTION_ID_COL).value)
                    if not question_id:
                        continue
                    sheet_stats["target_rows"] += 1
                    answer = model_answers[group_key][sheet_name].get(question_id, "")
                    if not answer:
                        sheet_stats["missing_model"] += 1
                        continue
                    counts = people_counts[group_key].get(question_id, Counter())
                    if not counts:
                        sheet_stats["missing_people"] += 1
                        continue
                    score = jsd_from_counter_and_answer(counts, answer)
                    if score is None:
                        continue
                    worksheet.cell(row=row_idx, column=jsd_col, value=score)
                    scores.append(score)
                    sheet_stats["written"] += 1
                if scores:
                    sheet_stats["mean_jsd"] = round(sum(scores) / len(scores), 6)
        workbook.save(field_path)
        return stats
    finally:
        workbook.close()


def print_stats(field_path: Path, stats: dict[str, dict[str, dict[str, int | float]]], write: bool) -> None:
    action_key = "written" if write else "calculated"
    print(f"\n{field_path.name}")
    for group_key, group in FIELD_GROUPS.items():
        print(f"  {group['label']}")
        for sheet_name in MODEL_SHEETS:
            sheet_stats = stats[group_key][sheet_name]
            print(
                f"    {sheet_name}: "
                f"target_rows={sheet_stats['target_rows']}, "
                f"{action_key}={sheet_stats[action_key]}, "
                f"missing_model={sheet_stats['missing_model']}, "
                f"missing_people={sheet_stats['missing_people']}, "
                f"mean_jsd={sheet_stats['mean_jsd']}"
            )


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Calculate JSD for Data quantifier and free-description quantifier rows in field workbooks."
    )
    parser.add_argument("--data", type=Path, default=DATA_RELATIVE_PATH)
    parser.add_argument(
        "--field",
        type=Path,
        action="append",
        default=None,
        help="Field workbook to calculate. Can be passed multiple times. Defaults to field-all/live/unlive.",
    )
    parser.add_argument("--write", action="store_true", help="Write per-row JSD scores into selected field workbooks.")
    args = parser.parse_args()

    people_counts = load_people_quantifier_counts(args.data)
    model_answers = load_model_quantifier_answers(args.data)
    field_paths = args.field if args.field else FIELD_RELATIVE_PATHS

    for field_path in field_paths:
        if args.write:
            stats = write_field_jsd(field_path, people_counts, model_answers)
        else:
            stats = calculate_field_jsd(field_path, people_counts, model_answers)
        print_stats(field_path, stats, args.write)

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
