from __future__ import annotations

import argparse
import math
import re
import string
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook


DATA_RELATIVE_PATH = Path("\u8f6c\u5f55\u6570\u636e") / "Data.xlsx"
FIELD_RELATIVE_PATHS = [
    Path("\u8f6c\u5f55\u6570\u636e") / "\u6307\u6807\u8ba1\u7b97" / "field-all.xlsx",
    Path("\u8f6c\u5f55\u6570\u636e") / "\u6307\u6807\u8ba1\u7b97" / "field-live.xlsx",
    Path("\u8f6c\u5f55\u6570\u636e") / "\u6307\u6807\u8ba1\u7b97" / "field-unlive.xlsx",
]

TASK_NOUN = "\u540d\u8bcd\u4efb\u52a1"
TASK_FREE_DESCRIPTION = "\u81ea\u7531\u63cf\u8ff0\u4efb\u52a1"
FIELD_NOUN_GROUPS = {
    "data_noun": {
        "label": "Data \u540d\u8bcd\u4efb\u52a1",
        "data_task": TASK_NOUN,
        "qid_prefixes": ("rd_", "rds-"),
        "field_signal_col": 2,
    },
    "free_description_quantifier": {
        "label": "Data \u81ea\u7531\u63cf\u8ff0-\u91cf\u8bcd\u4efb\u52a1",
        "data_task": TASK_FREE_DESCRIPTION,
        "qid_prefixes": ("rdf-",),
        "field_signal_col": 1,
    },
    "free_description_noun": {
        "label": "Data \u81ea\u7531\u63cf\u8ff0-\u540d\u8bcd\u4efb\u52a1",
        "data_task": TASK_FREE_DESCRIPTION,
        "qid_prefixes": ("rdf-",),
        "field_signal_col": 2,
    },
}
MODEL_SHEETS = [
    "glm",
    "gpt",
    "gemini",
    "qwen3.6-plus",
    "claude",
    "llama-4",
    "qwen3-vl-32b-instruct",
    "qwen2.5-vl-72b-instruct",
]

QUESTION_RE = re.compile(r"(\d+)\s*$")
PUNCTUATION_TABLE = str.maketrans(
    "",
    "",
    string.punctuation + "\u3002\uff0c\uff1f\uff01\uff1a\uff1b\u201c\u201d\u2018\u2019\u3001\uff08\uff09\u3010\u3011\u300a\u300b",
)


def clean_cell(value: object) -> str:
    if value is None:
        return ""
    return str(value).replace("_x000d_", "").replace("\r", "").replace("\n", "").strip()


def question_num(question_id: str) -> int | None:
    match = QUESTION_RE.search(clean_cell(question_id))
    return int(match.group(1)) if match else None


def has_prefix(question_id: str, prefixes: tuple[str, ...]) -> bool:
    return any(question_id.startswith(prefix) for prefix in prefixes)


def tokenize_zh_chars(text: str) -> list[str]:
    normalized = clean_cell(text).translate(PUNCTUATION_TABLE)
    normalized = re.sub(r"\s+", "", normalized)
    return list(normalized)


def ngrams(tokens: list[str], n: int) -> Counter[tuple[str, ...]]:
    if len(tokens) < n:
        return Counter()
    return Counter(tuple(tokens[i : i + n]) for i in range(len(tokens) - n + 1))


def modified_precision_counts(
    candidate: list[str],
    references: list[list[str]],
    n: int,
) -> tuple[int, int]:
    candidate_ngrams = ngrams(candidate, n)
    total = sum(candidate_ngrams.values())
    if total == 0:
        return 0, 0

    max_ref_counts: Counter[tuple[str, ...]] = Counter()
    for ref in references:
        max_ref_counts |= ngrams(ref, n)
    clipped = sum(min(count, max_ref_counts[gram]) for gram, count in candidate_ngrams.items())
    return clipped, total


def closest_reference_length(candidate_len: int, reference_lengths: list[int]) -> int:
    return min(reference_lengths, key=lambda length: (abs(length - candidate_len), length))


def sentence_bleu_char(candidate_text: str, reference_texts: list[str], max_n: int = 4) -> float:
    candidate = tokenize_zh_chars(candidate_text)
    references = [tokens for text in reference_texts if (tokens := tokenize_zh_chars(text))]
    if not candidate or not references:
        return 0.0

    candidate_len = len(candidate)
    reference_len = closest_reference_length(candidate_len, [len(ref) for ref in references])
    brevity_penalty = 1.0 if candidate_len > reference_len else math.exp(1 - reference_len / candidate_len)

    log_precision_sum = 0.0
    for n in range(1, max_n + 1):
        clipped, total = modified_precision_counts(candidate, references, n)
        if total == 0:
            # Add-one smoothing keeps short Chinese sentences from collapsing to 0.
            precision = 1.0
        else:
            precision = (clipped + 1) / (total + 1)
        log_precision_sum += math.log(precision)

    return round(brevity_penalty * math.exp(log_precision_sum / max_n), 6)


def corpus_bleu_char(
    candidate_references: list[tuple[str, list[str]]],
    max_n: int = 4,
) -> float:
    clipped_totals = [0] * max_n
    ngram_totals = [0] * max_n
    candidate_length = 0
    reference_length = 0

    for candidate_text, reference_texts in candidate_references:
        candidate = tokenize_zh_chars(candidate_text)
        references = [tokens for text in reference_texts if (tokens := tokenize_zh_chars(text))]
        if not references:
            continue

        candidate_length += len(candidate)
        reference_length += closest_reference_length(
            candidate_len=len(candidate),
            reference_lengths=[len(ref) for ref in references],
        )

        for n in range(1, max_n + 1):
            clipped, total = modified_precision_counts(candidate, references, n)
            clipped_totals[n - 1] += clipped
            ngram_totals[n - 1] += total

    if candidate_length == 0 or reference_length == 0:
        return 0.0

    log_precision_sum = 0.0
    for clipped, total in zip(clipped_totals, ngram_totals):
        if total == 0:
            precision = 1.0
        elif clipped == 0:
            return 0.0
        else:
            precision = clipped / total
        log_precision_sum += math.log(precision)

    brevity_penalty = 1.0 if candidate_length > reference_length else math.exp(1 - reference_length / candidate_length)
    return round(brevity_penalty * math.exp(log_precision_sum / max_n), 6)


def load_people_references_by_task(data_path: Path) -> dict[str, dict[str, list[str]]]:
    workbook = load_workbook(data_path, read_only=True, data_only=True)
    try:
        refs: dict[str, dict[str, list[str]]] = {
            group_key: defaultdict(list) for group_key in FIELD_NOUN_GROUPS
        }
        for row in workbook["people"].iter_rows(min_row=2, values_only=True):
            question_id = clean_cell(row[1])
            task = clean_cell(row[3])
            answer_text = clean_cell(row[2])
            if not answer_text:
                continue

            for group_key, group in FIELD_NOUN_GROUPS.items():
                if task == group["data_task"] and has_prefix(question_id, group["qid_prefixes"]):
                    refs[group_key][question_id].append(answer_text)
        return {group_key: dict(group_refs) for group_key, group_refs in refs.items()}
    finally:
        workbook.close()


def load_model_answers_by_task(data_path: Path) -> dict[str, dict[str, dict[str, str]]]:
    workbook = load_workbook(data_path, read_only=True, data_only=True)
    try:
        answers: dict[str, dict[str, dict[str, str]]] = {
            group_key: {sheet_name: {} for sheet_name in MODEL_SHEETS}
            for group_key in FIELD_NOUN_GROUPS
        }
        for sheet_name in MODEL_SHEETS:
            for row in workbook[sheet_name].iter_rows(min_row=2, values_only=True):
                question_id = clean_cell(row[1])
                task = clean_cell(row[3])
                answer_text = clean_cell(row[2])
                for group_key, group in FIELD_NOUN_GROUPS.items():
                    if task == group["data_task"] and has_prefix(question_id, group["qid_prefixes"]):
                        answers[group_key][sheet_name][question_id] = answer_text
        return answers
    finally:
        workbook.close()


def extract_field_noun_qids(field_path: Path) -> dict[str, dict[str, set[str]]]:
    workbook = load_workbook(field_path, read_only=True, data_only=True)
    try:
        qids_by_group: dict[str, dict[str, set[str]]] = {
            group_key: {sheet_name: set() for sheet_name in MODEL_SHEETS}
            for group_key in FIELD_NOUN_GROUPS
        }
        for sheet_name in MODEL_SHEETS:
            worksheet = workbook[sheet_name]
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                question_id = clean_cell(row[0])

                for group_key, group in FIELD_NOUN_GROUPS.items():
                    signal_value = clean_cell(row[group["field_signal_col"]])
                    if not signal_value:
                        continue
                    if has_prefix(question_id, group["qid_prefixes"]):
                        qids_by_group[group_key][sheet_name].add(question_id)
        return qids_by_group
    finally:
        workbook.close()


def compute_model_bleu_for_qids(
    references: dict[str, list[str]],
    answers_by_sheet: dict[str, dict[str, str]],
    qids_by_sheet: dict[str, set[str]],
) -> tuple[dict[str, dict[str, float]], dict[str, float]]:
    results: dict[str, dict[str, float]] = {}
    corpus_results: dict[str, float] = {}
    for sheet_name in MODEL_SHEETS:
        sheet_results: dict[str, float] = {}
        corpus_inputs: list[tuple[str, list[str]]] = []
        sheet_answers = answers_by_sheet[sheet_name]
        for question_id in sorted(qids_by_sheet[sheet_name], key=lambda qid: (question_num(qid) or 0, qid)):
            if question_id not in sheet_answers:
                continue
            question_references = references.get(question_id, [])
            answer_text = sheet_answers[question_id]
            sheet_results[question_id] = sentence_bleu_char(
                answer_text,
                question_references,
            )
            corpus_inputs.append((answer_text, question_references))
        results[sheet_name] = sheet_results
        corpus_results[sheet_name] = corpus_bleu_char(corpus_inputs)
    return results, corpus_results


def compute_field_bleu(
    field_path: Path,
    references_by_task: dict[str, dict[str, list[str]]],
    answers_by_task: dict[str, dict[str, dict[str, str]]],
) -> dict[str, tuple[dict[str, dict[str, float]], dict[str, float]]]:
    qids_by_group = extract_field_noun_qids(field_path)
    return {
        group_key: compute_model_bleu_for_qids(
            references=references_by_task[group_key],
            answers_by_sheet=answers_by_task[group_key],
            qids_by_sheet=qids_by_group[group_key],
        )
        for group_key in FIELD_NOUN_GROUPS
    }


def write_noun_bleu(
    field_path: Path,
    bleu_by_group: dict[str, dict[str, dict[str, float]]],
) -> dict[str, dict[str, int]]:
    workbook = load_workbook(field_path)
    try:
        write_counts: dict[str, dict[str, int]] = {
            group_key: {sheet_name: 0 for sheet_name in MODEL_SHEETS}
            for group_key in FIELD_NOUN_GROUPS
        }
        for sheet_name in MODEL_SHEETS:
            worksheet = workbook[sheet_name]
            header = [worksheet.cell(row=1, column=col).value for col in range(1, worksheet.max_column + 1)]
            try:
                bleu_col = header.index("Bleu") + 1
            except ValueError as exc:
                raise ValueError(f"{sheet_name} sheet has no Bleu column") from exc

            for row_idx in range(2, worksheet.max_row + 1):
                question_id = clean_cell(worksheet.cell(row=row_idx, column=1).value)

                for group_key, group in FIELD_NOUN_GROUPS.items():
                    signal_value = clean_cell(
                        worksheet.cell(row=row_idx, column=group["field_signal_col"] + 1).value
                    )
                    if not signal_value:
                        continue
                    if not has_prefix(question_id, group["qid_prefixes"]):
                        continue
                    sheet_bleu = bleu_by_group[group_key][sheet_name]
                    if question_id in sheet_bleu:
                        worksheet.cell(row=row_idx, column=bleu_col, value=sheet_bleu[question_id])
                        write_counts[group_key][sheet_name] += 1
        workbook.save(field_path)
        return write_counts
    finally:
        workbook.close()


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Calculate character-level mean and corpus BLEU for noun-task rows in field workbooks."
    )
    parser.add_argument("--data", type=Path, default=DATA_RELATIVE_PATH)
    parser.add_argument(
        "--field",
        type=Path,
        action="append",
        default=None,
        help="Field workbook to calculate. Can be passed multiple times. Defaults to field-all/live/unlive.",
    )
    parser.add_argument("--write", action="store_true", help="Write per-row sentence BLEU scores into selected field workbooks.")
    args = parser.parse_args()

    references_by_task = load_people_references_by_task(args.data)
    answers_by_task = load_model_answers_by_task(args.data)
    field_paths = args.field if args.field else FIELD_RELATIVE_PATHS

    for field_path in field_paths:
        field_results = compute_field_bleu(field_path, references_by_task, answers_by_task)
        print(f"\n{field_path.name}")

        for group_key, group in FIELD_NOUN_GROUPS.items():
            bleu_by_sheet, corpus_bleu_by_sheet = field_results[group_key]
            print(f"  {group['label']}")
            for sheet_name in MODEL_SHEETS:
                scores = list(bleu_by_sheet[sheet_name].values())
                average = round(sum(scores) / len(scores), 6) if scores else 0.0
                corpus_bleu = corpus_bleu_by_sheet[sheet_name]
                print(f"    {sheet_name}: rows={len(scores)}, mean_bleu={average}, corpus_bleu={corpus_bleu}")

        if args.write:
            write_counts = write_noun_bleu(
                field_path,
                {group_key: field_results[group_key][0] for group_key in FIELD_NOUN_GROUPS},
            )
            print("  written:", write_counts)

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
