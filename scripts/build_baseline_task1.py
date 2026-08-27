from __future__ import annotations

import argparse
import math
import re
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font


INPUT_RELATIVE_PATH = (
    Path("\u8f6c\u5f55\u6570\u636e")
    / "\u6307\u6807\u8ba1\u7b97"
    / "field-all-final.xlsx"
)
OUTPUT_RELATIVE_PATH = (
    Path("\u8f6c\u5f55\u6570\u636e")
    / "\u6307\u6807\u8ba1\u7b97"
    / "baseline-task1.xlsx"
)

DEFAULT_SHEET = "gpt"
TASK1_START_ROW = 2
TASK1_END_ROW = 181

QUESTION_ID_COL = 1
PEOPLE_DISTRIBUTION_COL = 4
PEOPLE_MOST_TYPE_COL = 8


@dataclass
class DetailRow:
    question_id: str
    people_distribution_text: str
    baseline_classifier: str
    baseline_classifier_type: str
    people_most_type: str
    rr: float
    right_wrong: int
    jsd: float

    def values(self) -> list[Any]:
        return [
            self.question_id,
            self.people_distribution_text,
            self.baseline_classifier,
            self.baseline_classifier_type,
            self.people_most_type,
            self.rr,
            self.right_wrong,
            self.jsd,
        ]


@dataclass
class SkippedRow:
    question_id: str
    row_number: int
    reason: str

    def values(self) -> list[Any]:
        return [self.question_id, self.row_number, self.reason]


def clean_cell(value: object) -> str:
    if value is None:
        return ""
    return str(value).replace("_x000d_", "").replace("\r", "").replace("\n", "").strip()


def parse_distribution_text(value: object) -> Counter[str]:
    text = clean_cell(value)
    counts: Counter[str] = Counter()
    if not text:
        return counts

    for item in re.split(r"[;\uff1b]", text):
        item = clean_cell(item)
        if not item:
            continue
        if "\uff1a" in item:
            key, raw_count = item.rsplit("\uff1a", 1)
        elif ":" in item:
            key, raw_count = item.rsplit(":", 1)
        else:
            continue
        key = clean_cell(key)
        raw_count = clean_cell(raw_count)
        if not key:
            continue
        try:
            count = int(float(raw_count))
        except ValueError:
            continue
        if count > 0:
            counts[key] += count
    return counts


def highest_frequency_first(counts: Counter[str]) -> str:
    if not counts:
        return ""
    max_count = max(counts.values())
    for key, count in counts.items():
        if count == max_count:
            return key
    return ""


def reciprocal_rank(counts: Counter[str], answer: str) -> float | None:
    answer = clean_cell(answer)
    if not counts or not answer:
        return None
    if answer not in counts:
        return 0.0

    target_count = counts[answer]
    rank = 1
    previous_count: int | None = None
    for _, count in counts.most_common():
        if previous_count is None:
            previous_count = count
        elif count != previous_count:
            rank += 1
            previous_count = count
        if count == target_count:
            return round(1 / rank, 6)
    return 0.0


def normalize_distribution(counts: Counter[str]) -> dict[str, float]:
    total = sum(counts.values())
    if total <= 0:
        return {}
    return {key: value / total for key, value in counts.items()}


def kl_divergence(p_dist: dict[str, float], q_dist: dict[str, float]) -> float:
    value = 0.0
    for key, p_value in p_dist.items():
        if p_value <= 0:
            continue
        q_value = q_dist.get(key, 0.0)
        if q_value <= 0:
            continue
        value += p_value * math.log2(p_value / q_value)
    return value


def jsd_from_counter_and_answer(counts: Counter[str], answer: str) -> float | None:
    answer = clean_cell(answer)
    if not counts or not answer:
        return None
    people_dist = normalize_distribution(counts)
    baseline_dist = {answer: 1.0}
    support = set(people_dist) | set(baseline_dist)
    midpoint = {
        key: (people_dist.get(key, 0.0) + baseline_dist.get(key, 0.0)) / 2
        for key in support
    }
    jsd = 0.5 * kl_divergence(people_dist, midpoint) + 0.5 * kl_divergence(baseline_dist, midpoint)
    return round(jsd, 6)


def mean(values: list[float]) -> float | None:
    if not values:
        return None
    return round(sum(values) / len(values), 6)


def build_rows(input_path: Path, sheet_name: str) -> tuple[list[DetailRow], list[SkippedRow]]:
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet_name]
        detail_rows: list[DetailRow] = []
        skipped_rows: list[SkippedRow] = []

        for row_number in range(TASK1_START_ROW, TASK1_END_ROW + 1):
            question_id = clean_cell(worksheet.cell(row=row_number, column=QUESTION_ID_COL).value)
            distribution_text = clean_cell(worksheet.cell(row=row_number, column=PEOPLE_DISTRIBUTION_COL).value)
            people_most_type = clean_cell(worksheet.cell(row=row_number, column=PEOPLE_MOST_TYPE_COL).value)
            counts = parse_distribution_text(distribution_text)
            if not counts:
                skipped_rows.append(
                    SkippedRow(
                        question_id=question_id,
                        row_number=row_number,
                        reason="empty or unparsable people_answer_calssifer",
                    )
                )
                continue

            baseline_classifier = highest_frequency_first(counts)
            baseline_classifier_type = people_most_type
            rr = reciprocal_rank(counts, baseline_classifier)
            jsd = jsd_from_counter_and_answer(counts, baseline_classifier)
            if rr is None or jsd is None:
                skipped_rows.append(
                    SkippedRow(
                        question_id=question_id,
                        row_number=row_number,
                        reason="failed to calculate RR or JSD",
                    )
                )
                continue

            right_wrong = 1 if baseline_classifier_type == people_most_type else 0
            detail_rows.append(
                DetailRow(
                    question_id=question_id,
                    people_distribution_text=distribution_text,
                    baseline_classifier=baseline_classifier,
                    baseline_classifier_type=baseline_classifier_type,
                    people_most_type=people_most_type,
                    rr=rr,
                    right_wrong=right_wrong,
                    jsd=jsd,
                )
            )
        return detail_rows, skipped_rows
    finally:
        workbook.close()


def write_output(output_path: Path, detail_rows: list[DetailRow], skipped_rows: list[SkippedRow]) -> None:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "summary"
    detail = workbook.create_sheet("detail")
    skipped = workbook.create_sheet("skipped")

    summary_headers = ["baseline_name", "valid_items", "skipped_items", "MRR", "accuracy", "JSD"]
    detail_headers = [
        "question_id",
        "people_answer_calssifer",
        "baseline_answer_classifier",
        "baseline_answer_classifier_type",
        "people_answer__classifer_most",
        "RR",
        "right_wrong",
        "JSD",
    ]
    skipped_headers = ["question_id", "row", "reason"]

    summary.append(summary_headers)
    rr_values = [row.rr for row in detail_rows]
    right_wrong_values = [row.right_wrong for row in detail_rows]
    jsd_values = [row.jsd for row in detail_rows]
    summary.append(
        [
            "Human-Most Classifier",
            len(detail_rows),
            len(skipped_rows),
            mean(rr_values),
            mean(right_wrong_values),
            mean(jsd_values),
        ]
    )

    detail.append(detail_headers)
    for row in detail_rows:
        detail.append(row.values())

    skipped.append(skipped_headers)
    for row in skipped_rows:
        skipped.append(row.values())

    for worksheet in (summary, detail, skipped):
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
        worksheet.freeze_panes = "A2"
        for column_cells in worksheet.columns:
            max_length = max(len(clean_cell(cell.value)) for cell in column_cells)
            worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 10), 48)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def main() -> int:
    parser = argparse.ArgumentParser(description="Build Task 1 human-most classifier baseline workbook.")
    parser.add_argument("--input", type=Path, default=INPUT_RELATIVE_PATH)
    parser.add_argument("--output", type=Path, default=OUTPUT_RELATIVE_PATH)
    parser.add_argument("--sheet", default=DEFAULT_SHEET)
    args = parser.parse_args()

    detail_rows, skipped_rows = build_rows(args.input, args.sheet)
    if len(detail_rows) < TASK1_END_ROW - TASK1_START_ROW + 1:
        print(
            f"warning: valid rows={len(detail_rows)}, expected={TASK1_END_ROW - TASK1_START_ROW + 1}, "
            f"skipped={len(skipped_rows)}"
        )
        for row in skipped_rows:
            print(f"  skipped row={row.row_number}, question_id={row.question_id}, reason={row.reason}")

    write_output(args.output, detail_rows, skipped_rows)
    print(f"written: {args.output}")
    print(
        f"summary: valid_items={len(detail_rows)}, skipped_items={len(skipped_rows)}, "
        f"MRR={mean([row.rr for row in detail_rows])}, "
        f"accuracy={mean([row.right_wrong for row in detail_rows])}, "
        f"JSD={mean([row.jsd for row in detail_rows])}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
