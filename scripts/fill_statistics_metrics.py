from __future__ import annotations

import argparse
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


STATS_RELATIVE_PATH = Path("\u8f6c\u5f55\u6570\u636e") / "\u6307\u6807\u8ba1\u7b97" / "\u7edf\u8ba1\u6307\u6807.xlsx"
FIELD_PATHS = {
    "\u6709\u751f\u547d\u6027": Path("\u8f6c\u5f55\u6570\u636e") / "\u6307\u6807\u8ba1\u7b97" / "field-live.xlsx",
    "\u65e0\u751f\u547d\u6027": Path("\u8f6c\u5f55\u6570\u636e") / "\u6307\u6807\u8ba1\u7b97" / "field-unlive.xlsx",
    "\u603b\u4f53": Path("\u8f6c\u5f55\u6570\u636e") / "\u6307\u6807\u8ba1\u7b97" / "field-all.xlsx",
}

MODEL_ROW_TO_FIELD_SHEET = {
    "glm": "glm",
    "gpt": "gpt",
    "gemini": "gemini",
    "qwen3.6-plus": "qwen3.6-plus",
    "claude": "claude",
    "llama-4": "llama-4",
    "qwen3-vl": "qwen3-vl-32b-instruct",
    "qwne2.5-vl": "qwen2.5-vl-72b-instruct",
}

TASK_LABELS = {
    "\u91cf\u8bcd": "data_quantifier",
    "\u540d\u8bcd": "data_noun",
    "\u81ea\u7531": "free_quantifier",
    "\u81ea\u7531\u91cf\u8bcd": "free_quantifier",
    "\u81ea\u7531\u540d\u8bcd": "free_noun",
    "\u65e0\u56fe\u91cf\u8bcd": "standalone_quantifier",
}

FIELD_RANGES = {
    "field-live.xlsx": {
        "data_quantifier": (2, 91),
        "data_noun": (92, 151),
        "free_quantifier": (152, 181),
        "free_noun": (182, 211),
        "standalone_quantifier": (212, 301),
    },
    "field-unlive.xlsx": {
        "data_quantifier": (2, 91),
        "data_noun": (92, 151),
        "free_quantifier": (152, 181),
        "free_noun": (182, 211),
        "standalone_quantifier": (212, 301),
    },
    "field-all.xlsx": {
        "data_quantifier": (2, 181),
        "data_noun": (182, 301),
        "free_quantifier": (302, 361),
        "free_noun": (362, 421),
        "standalone_quantifier": (422, 601),
    },
}

ACCURACY_COL = 11
BLEU_COL = 12
MRR_COL = 13
JSD_COL = 14


def clean_cell(value: object) -> str:
    if value is None:
        return ""
    return str(value).replace("_x000d_", "").replace("\r", "").replace("\n", "").strip()


def numeric(value: object) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def first_numeric_in_range(worksheet, start_row: int, end_row: int, column: int) -> float | None:
    for row_idx in range(start_row, end_row + 1):
        value = numeric(worksheet.cell(row=row_idx, column=column).value)
        if value is not None:
            return value
    return None


def mean_numeric_in_range(worksheet, start_row: int, end_row: int, column: int) -> float | None:
    values = [
        value
        for row_idx in range(start_row, end_row + 1)
        if (value := numeric(worksheet.cell(row=row_idx, column=column).value)) is not None
    ]
    if not values:
        return None
    return round(sum(values) / len(values), 6)


def load_field_metrics() -> dict[str, dict[str, dict[str, dict[str, float | None]]]]:
    metrics: dict[str, dict[str, dict[str, dict[str, float | None]]]] = {}
    for scope_label, path in FIELD_PATHS.items():
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            ranges = FIELD_RANGES[path.name]
            metrics[scope_label] = {}
            for model_name in MODEL_ROW_TO_FIELD_SHEET.values():
                worksheet = workbook[model_name]
                metrics[scope_label][model_name] = {}
                for task_key, (start_row, end_row) in ranges.items():
                    metrics[scope_label][model_name][task_key] = {
                        "mrr": first_numeric_in_range(worksheet, start_row, end_row, MRR_COL),
                        "accuracy": first_numeric_in_range(worksheet, start_row, end_row, ACCURACY_COL),
                        "bleu": mean_numeric_in_range(worksheet, start_row, end_row, BLEU_COL),
                        "jsd": mean_numeric_in_range(worksheet, start_row, end_row, JSD_COL),
                    }
        finally:
            workbook.close()
    return metrics


def parse_header(header: str) -> tuple[str, str, str] | None:
    parts = clean_cell(header).split("-")
    if len(parts) < 3:
        return None
    scope, task, metric = parts[0], parts[1], parts[2].lower()
    if scope not in FIELD_PATHS or task not in TASK_LABELS:
        return None
    if metric not in {"mrr", "accuracy", "bleu", "jsd"}:
        return None
    return scope, TASK_LABELS[task], metric


def fill_statistics(
    stats_path: Path,
    metrics: dict[str, dict[str, dict[str, dict[str, float | None]]]],
    write: bool,
) -> tuple[int, list[tuple[str, str, str]]]:
    workbook = load_workbook(stats_path)
    try:
        worksheet = workbook[workbook.sheetnames[0]]
        headers = [worksheet.cell(row=1, column=col).value for col in range(1, worksheet.max_column + 1)]
        updates = 0
        skipped: list[tuple[str, str, str]] = []

        for row_idx in range(2, worksheet.max_row + 1):
            stats_model_name = clean_cell(worksheet.cell(row=row_idx, column=1).value)
            if not stats_model_name:
                continue
            field_model_name = MODEL_ROW_TO_FIELD_SHEET.get(stats_model_name)
            if not field_model_name:
                skipped.append((stats_model_name, "", "unknown model"))
                continue

            for col_idx, header in enumerate(headers, start=1):
                parsed = parse_header(str(header) if header is not None else "")
                if parsed is None:
                    continue
                scope, task_key, metric_key = parsed
                value = metrics[scope][field_model_name][task_key].get(metric_key)
                if value is None:
                    skipped.append((stats_model_name, clean_cell(header), "missing value"))
                    continue
                if write:
                    worksheet.cell(row=row_idx, column=col_idx, value=value)
                updates += 1

        if write:
            workbook.save(stats_path)
        return updates, skipped
    finally:
        workbook.close()


def main() -> int:
    parser = argparse.ArgumentParser(description="Fill non-BLEU summary metrics into 统计指标.xlsx from field workbooks.")
    parser.add_argument("--stats", type=Path, default=STATS_RELATIVE_PATH)
    parser.add_argument("--write", action="store_true", help="Write values into 统计指标.xlsx. Without this, only dry-run.")
    args = parser.parse_args()

    metrics = load_field_metrics()
    updates, skipped = fill_statistics(args.stats, metrics, args.write)
    print(f"{'written' if args.write else 'dry_run'} updates={updates}")
    if skipped:
        print("skipped:")
        for model_name, header, reason in skipped:
            print(f"  {model_name} | {header} | {reason}")
    else:
        print("skipped=0")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
