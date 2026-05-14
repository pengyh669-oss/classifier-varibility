from __future__ import annotations

import argparse
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook


ROOT_DIR = Path(__file__).resolve().parents[1]
ANSWER_DIR = ROOT_DIR / "LLM_answer_new"
LOG_DIR = ROOT_DIR / "LLM_answer_new_logs"
DATA_DIR = ROOT_DIR / "\u8f6c\u5f55\u6570\u636e"

MATERIAL_PATH = DATA_DIR / "\u5b9e\u9a8c\u6750\u65991.xlsx"
TARGET_PATH = DATA_DIR / "LLM\u8f6c\u5f55-new.xlsx"
DATA_PATH = DATA_DIR / "Data.xlsx"
DEFAULT_REPORT_PATH = ROOT_DIR / "outputs" / "llm_transcription_new_report.xlsx"

COLUMNS = {
    "test_id": "test_id",
    "question_id": "question_id",
    "answer_text": "\u56de\u7b54\u6587\u672c",
    "condition": "\u5b9e\u9a8c\u6761\u4ef6",
    "quantifier": "\u91cf\u8bcd\uff08\u5185\u5bb9\uff09",
    "noun": "\u540d\u8bcd\uff08\u5185\u5bb9\uff09",
    "quantifier_type": "\u91cf\u8bcd\u7c7b\u578b",
    "noun_type": "\u540d\u8bcd\u7c7b\u578b",
    "item_time": "\u5355\u9898\u65f6\u95f4",
    "total_time": "\u603b\u65f6\u95f4",
}

MODEL_CONFIGS = [
    {
        "file_model": "glm-4.6v-new",
        "sheet": "glm",
        "test_id": "glm-4.6v",
        "type_sheet": "glm",
    },
    {
        "file_model": "gpt-5.4-new",
        "sheet": "gpt",
        "test_id": "gpt-5.4",
        "type_sheet": "gpt",
    },
    {
        "file_model": "gemini-3.1-pro-preview-new",
        "sheet": "gemini",
        "test_id": "gemini-3.1-pro-preview",
        "type_sheet": "gemini",
    },
    {
        "file_model": "qwen3.6-plus-new",
        "sheet": "qwen3.6-plus",
        "test_id": "qwen3.6-plus",
        "type_sheet": "qwen3.6-plus",
    },
    {
        "file_model": "claude-sonnet-4-6-new",
        "sheet": "claude",
        "test_id": "claude-sonnet-4-6",
        "type_sheet": "claude",
    },
    {
        "file_model": "llama-4-scout-new",
        "sheet": "llama-4",
        "test_id": "llama-4-scout",
        "type_sheet": "llama-4",
    },
    {
        "file_model": "qwen3-vl-32b-instruct-new",
        "sheet": "qwen3-vl-32b-instruct",
        "test_id": "qwen3-vl-32b-instruct",
        "type_sheet": "qwen3-vl-32b-instruct",
    },
    {
        "file_model": "qwen2.5-vl-72b-instruct-new",
        "sheet": "qwen2.5-vl-72b-instruct",
        "test_id": "qwen2.5-vl-72b-instruct",
        "type_sheet": "qwen2.5-vl-72b-instruct",
    },
]

# This is only a parsing lexicon. It is intentionally local and does not come
# from annotated transcription workbooks.
QUANTIFIERS = (
    "\u4e2a",
    "\u53ea",
    "\u4f4d",
    "\u540d",
    "\u8f86",
    "\u67b6",
    "\u8258",
    "\u682a",
    "\u68f5",
    "\u6735",
    "\u4ef6",
    "\u5f20",
    "\u6761",
    "\u628a",
    "\u95f4",
    "\u5ea7",
    "\u5bb6",
    "\u53f0",
    "\u672c",
    "\u5339",
    "\u5934",
    "\u53cc",
    "\u7247",
    "\u5757",
    "\u74f6",
    "\u676f",
    "\u7897",
    "\u76d2",
    "\u652f",
    "\u6839",
    "\u9876",
    "\u6247",
    "\u9762",
    "\u5c01",
    "\u5e45",
    "\u9897",
    "\u7c92",
    "\u76cf",
    "\u4efd",
    "\u53e3",
    "\u573a",
    "\u5957",
    "\u5bf9",
    "\u675f",
    "\u680b",
    "\u6b3e",
    "\u76d8",
    "\u79cd",
    "\u7ec4",
    "\u7fa4",
    "\u9053",
    "\u6392",
    "\u90e8",
)

QUESTION_ID_RE = re.compile(r"(rd(?:s|u|d|f|n)?[-_]\d+)\s*$")
ANSWER_HEADER_RE = re.compile(r"^\[(\d+)/(\d+)]\s+\u6587\u4ef6:\s+(.+?)\s*$")
ANSWER_MARKER = "\u539f\u59cb\u8f93\u51fa:"
BLOCK_SEPARATOR_RE = re.compile(r"^-{10,}\s*$")
LOG_HEADER_RE = re.compile(
    r"^\[(\d+)/(\d+)]\s+"
    r"(result_data(?:_special|_free|_downWord|_upWord|_normalWord)?)"
    r"\s+\|\s+(\d+)\s*$"
)
LOG_TIME_RE = re.compile(r"->\s+(OK|FAIL)\s+\(([-+]?\d+(?:\.\d+)?)s\)")
IDENT_RE = re.compile(
    r"^LLM_[^_]+_(result_data(?:_special|_free|_downWord|_upWord|_normalWord)?)_"
    r"(\d+)_(rd(?:s|u|d|f|n)?[-_]\d+)$"
)
NUMERAL_CHARS = set("\u4e00\u4e24\u4e8c\u4e09\u56db\u4e94\u516d\u4e03\u516b\u4e5d\u5341")
TRIM_CHARS = " \t\r\n\u3000\u3002\uff01\uff1f!?，,\uff0c；;\uff1b：:\u3001\"'“”‘’（）()[]【】"


@dataclass
class AnswerRecord:
    index: int
    total: int
    identifier: str
    question_id: str
    source: str
    object_id: str
    answer_text: str


@dataclass
class LogRecord:
    index: int
    total: int
    source: str
    object_id: str
    status: str
    seconds: float


@dataclass
class OutputRow:
    test_id: str
    question_id: str
    answer_text: str
    condition: str
    quantifier: str
    noun: str
    quantifier_type: str
    noun_type: str
    item_time: float | None
    total_time: float


def clean_cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("_x000d_", "").replace("\r", "").replace("\n", "").strip()


def resolve_path(value: str | None, default: Path) -> Path:
    if not value:
        return default
    path = Path(value)
    return path if path.is_absolute() else ROOT_DIR / path


def header_map(worksheet) -> dict[str, int]:
    result: dict[str, int] = {}
    for cell in worksheet[1]:
        name = clean_cell(cell.value)
        if name and name not in result:
            result[name] = cell.column
    return result


def require_columns(worksheet, names: list[str]) -> dict[str, int]:
    mapping = header_map(worksheet)
    missing = [name for name in names if name not in mapping]
    if missing:
        raise ValueError(f"Sheet {worksheet.title} missing columns: {missing}")
    return {name: mapping[name] for name in names}


def load_material_conditions(path: Path) -> dict[str, str]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    cols = require_columns(ws, [COLUMNS["question_id"], COLUMNS["condition"]])
    result: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2):
        question_id = clean_cell(row[cols[COLUMNS["question_id"]] - 1].value)
        condition = clean_cell(row[cols[COLUMNS["condition"]] - 1].value)
        if question_id:
            result[question_id] = condition
    wb.close()
    return result


def quantifier_type_for(quantifier: str) -> str:
    if not quantifier:
        return ""
    return "\u666e\u901a\u91cf\u8bcd" if quantifier == "\u4e2a" else "\u7279\u6b8a\u91cf\u8bcd"


def load_noun_type_lookup(path: Path, sheet_names: list[str]) -> dict[tuple[str, str, str], str]:
    wb = load_workbook(path, read_only=True, data_only=True)
    aggregated: dict[tuple[str, str, str], Counter[str]] = defaultdict(Counter)
    required = [
        COLUMNS["question_id"],
        COLUMNS["quantifier"],
        COLUMNS["noun"],
        COLUMNS["noun_type"],
    ]
    for sheet_name in sheet_names:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Data.xlsx missing sheet: {sheet_name}")
        ws = wb[sheet_name]
        cols = require_columns(ws, required)
        lookup: dict[tuple[str, str, str], tuple[str, str]] = {}
        for row in ws.iter_rows(min_row=2):
            question_id = clean_cell(row[cols[COLUMNS["question_id"]] - 1].value)
            quantifier = clean_cell(row[cols[COLUMNS["quantifier"]] - 1].value)
            noun = clean_cell(row[cols[COLUMNS["noun"]] - 1].value)
            noun_type = clean_cell(row[cols[COLUMNS["noun_type"]] - 1].value)
            if question_id and noun_type:
                aggregated[(question_id, quantifier, noun)][noun_type] += 1
    wb.close()
    return {key: counts.most_common(1)[0][0] for key, counts in aggregated.items()}


def load_answer_fallback_lookup(path: Path, configs: list[dict[str, str]]) -> dict[str, dict[str, str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    lookups: dict[str, dict[str, str]] = {}
    required = [COLUMNS["question_id"], COLUMNS["answer_text"]]
    for config in configs:
        sheet_name = config["type_sheet"]
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Data.xlsx missing sheet: {sheet_name}")
        ws = wb[sheet_name]
        cols = require_columns(ws, required)
        lookup: dict[str, str] = {}
        for row in ws.iter_rows(min_row=2):
            question_id = clean_cell(row[cols[COLUMNS["question_id"]] - 1].value)
            answer_text = clean_cell(row[cols[COLUMNS["answer_text"]] - 1].value)
            if question_id and answer_text:
                lookup[question_id] = answer_text
        lookups[config["file_model"]] = lookup
    wb.close()
    return lookups


def parse_identifier(identifier: str) -> tuple[str, str, str]:
    match = IDENT_RE.match(identifier)
    if not match:
        qid_match = QUESTION_ID_RE.search(identifier)
        return "", "", qid_match.group(1) if qid_match else ""
    return match.group(1), match.group(2), match.group(3)


def parse_answer_file(path: Path) -> list[AnswerRecord]:
    records: list[AnswerRecord] = []
    current: dict[str, Any] | None = None
    answer_lines: list[str] = []

    def finish_current() -> None:
        nonlocal current, answer_lines
        if not current:
            return
        answer_text = "\n".join(answer_lines).strip()
        source, object_id, parsed_qid = parse_identifier(current["identifier"])
        question_id = parsed_qid or current["question_id"]
        records.append(
            AnswerRecord(
                index=current["index"],
                total=current["total"],
                identifier=current["identifier"],
                question_id=question_id,
                source=source,
                object_id=object_id,
                answer_text=answer_text,
            )
        )
        current = None
        answer_lines = []

    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.rstrip("\n")
        header_match = ANSWER_HEADER_RE.match(line.strip())
        if header_match:
            finish_current()
            identifier = header_match.group(3).strip()
            qid_match = QUESTION_ID_RE.search(identifier)
            current = {
                "index": int(header_match.group(1)),
                "total": int(header_match.group(2)),
                "identifier": identifier,
                "question_id": qid_match.group(1) if qid_match else "",
            }
            answer_lines = []
            continue

        if not current:
            continue

        if BLOCK_SEPARATOR_RE.match(line.strip()):
            finish_current()
            continue

        marker_pos = line.find(ANSWER_MARKER)
        if marker_pos >= 0:
            answer_lines.append(line[marker_pos + len(ANSWER_MARKER) :].strip())
        elif answer_lines:
            answer_lines.append(line)

    finish_current()
    return records


def parse_log_file(path: Path) -> list[LogRecord]:
    records: list[LogRecord] = []
    current: dict[str, Any] | None = None
    for line in path.read_text(encoding="utf-8").splitlines():
        header_match = LOG_HEADER_RE.match(line.strip())
        if header_match:
            current = {
                "index": int(header_match.group(1)),
                "total": int(header_match.group(2)),
                "source": header_match.group(3),
                "object_id": header_match.group(4),
            }
            continue

        time_match = LOG_TIME_RE.search(line)
        if time_match and current:
            records.append(
                LogRecord(
                    index=current["index"],
                    total=current["total"],
                    source=current["source"],
                    object_id=current["object_id"],
                    status=time_match.group(1),
                    seconds=float(time_match.group(2)),
                )
            )
            current = None
    return records


def normalize_answer_for_extraction(answer_text: str) -> str:
    return answer_text.replace("\\n", "\n").replace("\\r", "\n")


def trim_noun(text: str) -> str:
    value = text.strip(TRIM_CHARS)
    value = re.split(r"[\n\r\u3002\uff01\uff1f!?，,\uff0c；;\uff1b：:\u3001]", value, maxsplit=1)[0]
    return value.strip(TRIM_CHARS)


def extract_quantifier_noun(answer_text: str) -> tuple[str, str]:
    text = normalize_answer_for_extraction(answer_text)
    candidates: list[tuple[str, str]] = []
    quantifiers = sorted(set(QUANTIFIERS), key=len, reverse=True)

    for index, char in enumerate(text):
        if char not in NUMERAL_CHARS:
            continue
        rest = text[index + 1 :]
        for quantifier in quantifiers:
            if rest.startswith(quantifier):
                noun = trim_noun(rest[len(quantifier) :])
                if noun:
                    candidates.append((quantifier, noun))
                break

    return candidates[-1] if candidates else ("", "")


def build_rows_for_model(
    config: dict[str, str],
    material_conditions: dict[str, str],
    noun_type_lookup: dict[tuple[str, str, str], str],
    answer_fallback_lookup: dict[str, str],
    report_rows: list[dict[str, Any]],
) -> tuple[list[OutputRow], dict[str, Any]]:
    answer_path = ANSWER_DIR / f"LLM_{config['file_model']}.txt"
    log_path = LOG_DIR / f"{config['file_model']}.txt"
    if not answer_path.exists():
        raise FileNotFoundError(answer_path)
    if not log_path.exists():
        raise FileNotFoundError(log_path)

    answers = parse_answer_file(answer_path)
    logs = parse_log_file(log_path)
    log_by_index = {record.index: record for record in logs}
    total_time = round(sum(record.seconds for record in logs), 2)
    rows: list[OutputRow] = []

    duplicate_qids = len(answers) - len({record.question_id for record in answers})
    stats = {
        "model": config["file_model"],
        "sheet": config["sheet"],
        "answers": len(answers),
        "logs": len(logs),
        "unique_question_ids": len({record.question_id for record in answers}),
        "duplicate_question_ids": duplicate_qids,
        "missing_material": 0,
        "missing_extraction": 0,
        "missing_type": 0,
        "sequence_mismatch": 0,
        "answer_text_fallback": 0,
        "fail_logs": sum(1 for record in logs if record.status == "FAIL"),
        "total_time": total_time,
    }

    for answer in answers:
        condition = material_conditions.get(answer.question_id, "")
        if not condition:
            stats["missing_material"] += 1
            report_rows.append(
                issue(config, answer, "material_missing", "question_id not found in material")
            )

        log_record = log_by_index.get(answer.index)
        item_time: float | None = None
        if log_record:
            item_time = log_record.seconds
            if (
                log_record.source != answer.source
                or log_record.object_id != answer.object_id
                or log_record.total != answer.total
            ):
                stats["sequence_mismatch"] += 1
                report_rows.append(
                    issue(
                        config,
                        answer,
                        "txt_log_mismatch",
                        (
                            f"txt={answer.source}|{answer.object_id}|{answer.total}; "
                            f"log={log_record.source}|{log_record.object_id}|{log_record.total}"
                        ),
                    )
                )
        else:
            stats["sequence_mismatch"] += 1
            report_rows.append(issue(config, answer, "log_missing", "log index not found"))

        answer_text = answer.answer_text
        needs_answer_fallback = not answer_text.strip() or (
            log_record is not None and log_record.status == "FAIL"
        )
        if needs_answer_fallback:
            fallback_answer = answer_fallback_lookup.get(answer.question_id, "")
            if fallback_answer:
                answer_text = fallback_answer
                stats["answer_text_fallback"] += 1
                report_rows.append(
                    issue(
                        config,
                        answer,
                        "answer_text_fallback",
                        "blank or failed original answer filled from Data.xlsx corresponding model sheet",
                    )
                )

        quantifier, noun = extract_quantifier_noun(answer_text)
        if not quantifier or not noun:
            stats["missing_extraction"] += 1
            report_rows.append(
                issue(
                    config,
                    answer,
                    "extraction_missing",
                    "could not parse quantifier/noun from original answer",
                )
            )

        quantifier_type = quantifier_type_for(quantifier)
        noun_type = noun_type_lookup.get((answer.question_id, quantifier, noun), "")
        if (quantifier or noun) and not noun_type:
            stats["missing_type"] += 1
            report_rows.append(
                issue(
                    config,
                    answer,
                    "type_missing",
                    f"no Data.xlsx exact noun type match across model sheets for ({answer.question_id}, {quantifier}, {noun})",
                    quantifier=quantifier,
                    noun=noun,
                )
            )

        rows.append(
            OutputRow(
                test_id=config["test_id"],
                question_id=answer.question_id,
                answer_text=answer_text,
                condition=condition,
                quantifier=quantifier,
                noun=noun,
                quantifier_type=quantifier_type,
                noun_type=noun_type,
                item_time=item_time,
                total_time=total_time,
            )
        )

    return rows, stats


def issue(
    config: dict[str, str],
    answer: AnswerRecord,
    issue_type: str,
    detail: str,
    quantifier: str = "",
    noun: str = "",
) -> dict[str, Any]:
    return {
        "model": config["file_model"],
        "sheet": config["sheet"],
        "row_index": answer.index,
        "identifier": answer.identifier,
        "question_id": answer.question_id,
        "issue": issue_type,
        "detail": detail,
        "quantifier": quantifier,
        "noun": noun,
        "answer_text": answer.answer_text,
    }


def clear_sheet_data(worksheet) -> None:
    if worksheet.max_row >= 2:
        worksheet.delete_rows(2, worksheet.max_row - 1)


def write_rows_to_sheet(worksheet, rows: list[OutputRow]) -> None:
    cols = require_columns(worksheet, list(COLUMNS.values()))
    field_order = [
        ("test_id", "test_id"),
        ("question_id", "question_id"),
        ("answer_text", "answer_text"),
        ("condition", "condition"),
        ("quantifier", "quantifier"),
        ("noun", "noun"),
        ("quantifier_type", "quantifier_type"),
        ("noun_type", "noun_type"),
        ("item_time", "item_time"),
        ("total_time", "total_time"),
    ]
    for row_index, output_row in enumerate(rows, start=2):
        for attr_name, column_key in field_order:
            worksheet.cell(
                row=row_index,
                column=cols[COLUMNS[column_key]],
                value=getattr(output_row, attr_name),
            )


def save_report(path: Path, stats_rows: list[dict[str, Any]], issue_rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "summary"
    if stats_rows:
        headers = list(stats_rows[0].keys())
        summary_ws.append(headers)
        for row in stats_rows:
            summary_ws.append([row.get(header, "") for header in headers])
    else:
        summary_ws.append(["message"])
        summary_ws.append(["no stats"])

    issue_ws = wb.create_sheet("issues")
    if issue_rows:
        headers = list(issue_rows[0].keys())
        issue_ws.append(headers)
        for row in issue_rows:
            issue_ws.append([row.get(header, "") for header in headers])
    else:
        issue_ws.append(["message"])
        issue_ws.append(["no issues"])

    wb.save(path)


def process(output_path: Path, report_path: Path, dry_run: bool) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    material_conditions = load_material_conditions(MATERIAL_PATH)
    noun_type_lookup = load_noun_type_lookup(DATA_PATH, [config["type_sheet"] for config in MODEL_CONFIGS])
    answer_fallback_lookups = load_answer_fallback_lookup(DATA_PATH, MODEL_CONFIGS)
    report_rows: list[dict[str, Any]] = []
    stats_rows: list[dict[str, Any]] = []
    rows_by_sheet: dict[str, list[OutputRow]] = {}

    for config in MODEL_CONFIGS:
        rows, stats = build_rows_for_model(
            config=config,
            material_conditions=material_conditions,
            noun_type_lookup=noun_type_lookup,
            answer_fallback_lookup=answer_fallback_lookups[config["file_model"]],
            report_rows=report_rows,
        )
        rows_by_sheet[config["sheet"]] = rows
        stats_rows.append(stats)

    if not dry_run:
        wb = load_workbook(TARGET_PATH)
        for config in MODEL_CONFIGS:
            sheet_name = config["sheet"]
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"Target workbook missing sheet: {sheet_name}")
            ws = wb[sheet_name]
            clear_sheet_data(ws)
            write_rows_to_sheet(ws, rows_by_sheet[sheet_name])
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        save_report(report_path, stats_rows, report_rows)

    return stats_rows, report_rows


def print_summary(stats_rows: list[dict[str, Any]], report_rows: list[dict[str, Any]], dry_run: bool) -> None:
    mode = "DRY RUN" if dry_run else "WRITE"
    print(f"Mode: {mode}")
    print(
        "model | answers | logs | unique_qids | fail_logs | missing_material | "
        "missing_extraction | missing_type | sequence_mismatch | answer_text_fallback | total_time"
    )
    for row in stats_rows:
        print(
            f"{row['model']} | {row['answers']} | {row['logs']} | {row['unique_question_ids']} | "
            f"{row['fail_logs']} | {row['missing_material']} | {row['missing_extraction']} | "
            f"{row['missing_type']} | {row['sequence_mismatch']} | "
            f"{row['answer_text_fallback']} | {row['total_time']}"
        )
    print(f"Total output rows: {sum(row['answers'] for row in stats_rows)}")
    print(f"Report issue rows: {len(report_rows)}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Fill LLM transcription workbook from new LLM answer files.")
    parser.add_argument("--dry-run", action="store_true", help="Parse and validate without saving workbooks.")
    parser.add_argument("--output", help="Output xlsx path. Defaults to overwriting LLM\u8f6c\u5f55-new.xlsx.")
    parser.add_argument("--report", help="Report xlsx path. Defaults to outputs/llm_transcription_new_report.xlsx.")
    args = parser.parse_args()

    output_path = resolve_path(args.output, TARGET_PATH)
    report_path = resolve_path(args.report, DEFAULT_REPORT_PATH)
    stats_rows, report_rows = process(output_path=output_path, report_path=report_path, dry_run=args.dry_run)
    print_summary(stats_rows, report_rows, dry_run=args.dry_run)
    if not args.dry_run:
        print(f"Saved workbook: {output_path}")
        print(f"Saved report: {report_path}")


if __name__ == "__main__":
    main()
